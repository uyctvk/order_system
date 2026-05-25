from __future__ import annotations

import sqlite3
import secrets
import socket
from datetime import datetime
from pathlib import Path
from typing import Optional
import re
import json
import os
import time
import shutil
import io
import csv
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, jsonify, abort, flash, send_file, send_from_directory, session
import qrcode
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from werkzeug.security import check_password_hash, generate_password_hash

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "db" / "kakashi.db"
QR_DIR = BASE_DIR / "static" / "qr"
PDF_DIR = BASE_DIR / "receipts"
DAILY_REPORT_DIR = BASE_DIR / "daily_reports"
PHOTO_DIR = BASE_DIR / "photos"
MENU_TXT_PATH = BASE_DIR / "product_menu.txt"
CONFIG_PATH = BASE_DIR / "config.txt"
MENU_DISPLAY_TXT_PATH = BASE_DIR / "menu_display.txt"
MENU_TRANSLATIONS_TXT_PATH = BASE_DIR / "menu_translations.txt"
USER_TXT_PATH = BASE_DIR / "staff_users.txt"
STORE_NAME = "かかし 大牟田店"
TABLE_COUNT = 21
TEST_TABLE_NO = 999

APP_PORT = 5000


def load_config() -> dict:
    config = {}
    if not CONFIG_PATH.exists():
        return config
    for line in CONFIG_PATH.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not s or s.startswith("#") or "=" not in s:
            continue
        key, value = s.split("=", 1)
        config[key.strip()] = value.strip()
    return config


def _to_int(value: str | None, default: int, minimum: int | None = None, maximum: int | None = None) -> int:
    try:
        result = int(str(value or "").strip())
    except (TypeError, ValueError):
        result = default
    if minimum is not None:
        result = max(minimum, result)
    if maximum is not None:
        result = min(maximum, result)
    return result


def _to_bool(value: str | None, default: bool = True) -> bool:
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on", "y"}


def load_simple_kv_txt(path: Path) -> dict[str, str]:
    """KEY=VALUE 形式TXTを読む。#コメントと空行は無視する。"""
    result: dict[str, str] = {}
    if not path.exists():
        return result
    for line in path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not s or s.startswith("#") or "=" not in s:
            continue
        key, value = s.split("=", 1)
        result[key.strip().upper()] = value.strip()
    return result


def load_menu_display_settings() -> dict:
    """
    メニュー画面の表示固定値。
    お客様側で写真サイズ・文字サイズを変更させず、menu_display.txt で固定管理する。
    """
    raw = load_simple_kv_txt(MENU_DISPLAY_TXT_PATH)
    ratio = raw.get("PHOTO_RATIO", "free").strip().lower()
    if ratio not in {"free", "1:1", "4:3", "3:2", "16:9"}:
        ratio = "free"

    photo_width = _to_int(raw.get("PHOTO_WIDTH"), 132, 70, 260)
    photo_height = _to_int(raw.get("PHOTO_HEIGHT"), 132, 70, 280)
    if ratio == "1:1":
        photo_height = photo_width
    elif ratio == "4:3":
        photo_height = round(photo_width * 3 / 4)
    elif ratio == "3:2":
        photo_height = round(photo_width * 2 / 3)
    elif ratio == "16:9":
        photo_height = round(photo_width * 9 / 16)

    text_scale = _to_int(raw.get("TEXT_SCALE"), 100, 80, 150)
    popup_photo_height = _to_int(raw.get("POPUP_PHOTO_HEIGHT"), max(160, round(photo_height * 1.55)), 120, 420)
    default_language = raw.get("DEFAULT_LANGUAGE", "ja").strip().lower()
    if default_language not in {"ja", "zh", "ko", "en"}:
        default_language = "ja"
    object_fit = raw.get("PHOTO_OBJECT_FIT", "contain").strip().lower()
    if object_fit not in {"contain", "cover"}:
        object_fit = "contain"

    return {
        "photo_width": photo_width,
        "photo_height": photo_height,
        "popup_photo_height": popup_photo_height,
        "text_scale": text_scale,
        "text_scale_ratio": f"{text_scale / 100:.3f}",
        "photo_ratio": ratio,
        "photo_object_fit": object_fit,
        "default_language": default_language,
        "enable_language_select": _to_bool(raw.get("ENABLE_LANGUAGE_SELECT"), True),
    }


TRANSLATION_EXACT: dict[str, dict[str, str]] = {
    # ジャンル
    "御前": {"zh": "御膳", "ko": "정식", "en": "Set Meal"},
    "天丼": {"zh": "天妇罗盖饭", "ko": "텐동", "en": "Tempura Bowl"},
    "甘味": {"zh": "甜品", "ko": "디저트", "en": "Dessert"},
    "フェア": {"zh": "限定推荐", "ko": "페어 메뉴", "en": "Seasonal Fair"},
    "釜飯": {"zh": "釜饭", "ko": "가마솥밥", "en": "Kamameshi"},
    "懐石": {"zh": "会席料理", "ko": "가이세키", "en": "Kaiseki"},
    "松花堂": {"zh": "松花堂便当", "ko": "쇼카도", "en": "Shokado Bento"},
    "丼": {"zh": "盖饭", "ko": "덮밥", "en": "Rice Bowl"},
    "うなぎ": {"zh": "鳗鱼", "ko": "장어", "en": "Eel"},
    "麺": {"zh": "面类", "ko": "면류", "en": "Noodles"},

    # 商品名
    "鶏唐揚げ定食": {"zh": "炸鸡块定食", "ko": "닭튀김 정식", "en": "Fried Chicken Set Meal"},
    "大エビ天丼味噌汁セット": {"zh": "大虾天妇罗盖饭味噌汤套餐", "ko": "대새우 텐동 된장국 세트", "en": "Large Shrimp Tempura Bowl with Miso Soup"},
    "大エビ天丼面セット": {"zh": "大虾天妇罗盖饭面套餐", "ko": "대새우 텐동 면 세트", "en": "Large Shrimp Tempura Bowl with Noodles"},
    "白玉あんみつ+抹茶": {"zh": "白玉馅蜜＋抹茶", "ko": "시라타마 안미쓰＋말차", "en": "Shiratama Anmitsu + Matcha"},
    "バナナクリームあんみつ": {"zh": "香蕉奶油馅蜜", "ko": "바나나 크림 안미쓰", "en": "Banana Cream Anmitsu"},
    "白玉抹茶パフェ": {"zh": "白玉抹茶芭菲", "ko": "시라타마 말차 파르페", "en": "Shiratama Matcha Parfait"},
    "小倉バナナパフェ": {"zh": "小仓香蕉芭菲", "ko": "오구라 바나나 파르페", "en": "Ogura Banana Parfait"},
    "餅入りぜんざい": {"zh": "年糕红豆汤", "ko": "떡 단팥죽", "en": "Sweet Red Bean Soup with Mochi"},
    "白玉ぜんざい": {"zh": "白玉红豆汤", "ko": "시라타마 단팥죽", "en": "Sweet Red Bean Soup with Shiratama"},
    "手作りわらび餅＋抹茶": {"zh": "手作蕨饼＋抹茶", "ko": "수제 와라비모치＋말차", "en": "Handmade Warabi Mochi + Matcha"},
    "抹茶アイス": {"zh": "抹茶冰淇淋", "ko": "말차 아이스크림", "en": "Matcha Ice Cream"},
    "バニラアイス": {"zh": "香草冰淇淋", "ko": "바닐라 아이스크림", "en": "Vanilla Ice Cream"},
    "春の旬かご懐石": {"zh": "春季时令笼会席", "ko": "봄 제철 카고 가이세키", "en": "Spring Seasonal Basket Kaiseki"},
    "ヒレカツ鳥窯セット": {"zh": "里脊猪排鸡肉釜饭套餐", "ko": "히레카츠 닭 가마솥밥 세트", "en": "Pork Fillet Cutlet & Chicken Kamameshi Set"},
    "ヒレカツウナギ窯セット": {"zh": "里脊猪排鳗鱼釜饭套餐", "ko": "히레카츠 장어 가마솥밥 세트", "en": "Pork Fillet Cutlet & Eel Kamameshi Set"},
    "小町懐石": {"zh": "小町会席", "ko": "고마치 가이세키", "en": "Komachi Kaiseki"},
    "つばき": {"zh": "椿", "ko": "츠바키", "en": "Tsubaki"},
    "大海老天ざる御前そば": {"zh": "大虾天妇罗竹筛荞麦御膳", "ko": "대새우 튀김 자루소바 정식", "en": "Large Shrimp Tempura Zaru Soba Set"},
    "ミニヒレカツ丼": {"zh": "迷你里脊猪排盖饭", "ko": "미니 히레카츠 덮밥", "en": "Mini Pork Fillet Cutlet Bowl"},
    "うな重上セット": {"zh": "上等鳗鱼饭套餐", "ko": "특상 장어덮밥 세트", "en": "Premium Eel Rice Box Set"},
    "ざるそば": {"zh": "竹筛荞麦面", "ko": "자루소바", "en": "Cold Zaru Soba"},

    # 選択肢
    "変更なし": {"zh": "不变更", "ko": "변경 없음", "en": "No change"},
    "なし": {"zh": "无", "ko": "없음", "en": "None"},
    "ネギ抜き": {"zh": "不要葱", "ko": "파 빼기", "en": "No green onion"},
    "ねぎ抜き": {"zh": "不要葱", "ko": "파 빼기", "en": "No green onion"},
    "ワサビ抜き": {"zh": "不要芥末", "ko": "와사비 빼기", "en": "No wasabi"},
    "ミニ温うどん変更": {"zh": "改为迷你热乌冬", "ko": "미니 온우동으로 변경", "en": "Change to mini hot udon"},
    "ミニ温そば変更": {"zh": "改为迷你热荞麦面", "ko": "미니 온소바로 변경", "en": "Change to mini hot soba"},
    "温そば変更": {"zh": "改为热荞麦面", "ko": "온소바로 변경", "en": "Change to hot soba"},
    "ざるそば変更": {"zh": "改为竹筛荞麦面", "ko": "자루소바로 변경", "en": "Change to cold zaru soba"},
    "ざるうどん変更": {"zh": "改为竹筛乌冬", "ko": "자루우동으로 변경", "en": "Change to cold zaru udon"},
    "温うどん変更": {"zh": "改为热乌冬", "ko": "온우동으로 변경", "en": "Change to hot udon"},
    "ミニ温うどん": {"zh": "迷你热乌冬", "ko": "미니 온우동", "en": "Mini hot udon"},
    "ミニ温そば": {"zh": "迷你热荞麦面", "ko": "미니 온소바", "en": "Mini hot soba"},
    "ミニ冷うどん": {"zh": "迷你冷乌冬", "ko": "미니 냉우동", "en": "Mini cold udon"},
    "ミニ冷そば": {"zh": "迷你冷荞麦面", "ko": "미니 냉소바", "en": "Mini cold soba"},
    "温そば": {"zh": "热荞麦面", "ko": "온소바", "en": "Hot soba"},
    "温うどん": {"zh": "热乌冬", "ko": "온우동", "en": "Hot udon"},
    "ざるうどん": {"zh": "竹筛乌冬", "ko": "자루우동", "en": "Cold zaru udon"},
    "大盛": {"zh": "大份", "ko": "곱빼기", "en": "Large portion"},
    "普通盛": {"zh": "普通份", "ko": "보통", "en": "Regular portion"},
    "天ぷら盛り合わせ": {"zh": "天妇罗拼盘", "ko": "튀김 모둠", "en": "Assorted tempura"},
    "握り寿司3貫": {"zh": "握寿司3贯", "ko": "니기리 초밥 3개", "en": "3 pieces of nigiri sushi"},
}

TRANSLATION_TERMS: dict[str, dict[str, str]] = {
    "大海老": {"zh": "大虾", "ko": "대새우", "en": "large shrimp"},
    "大エビ": {"zh": "大虾", "ko": "대새우", "en": "large shrimp"},
    "エビ": {"zh": "虾", "ko": "새우", "en": "shrimp"},
    "海老": {"zh": "虾", "ko": "새우", "en": "shrimp"},
    "ヒレカツ": {"zh": "里脊猪排", "ko": "히레카츠", "en": "pork fillet cutlet"},
    "天丼": {"zh": "天妇罗盖饭", "ko": "텐동", "en": "tempura bowl"},
    "味噌汁": {"zh": "味噌汤", "ko": "된장국", "en": "miso soup"},
    "セット": {"zh": "套餐", "ko": "세트", "en": "set"},
    "定食": {"zh": "定食", "ko": "정식", "en": "set meal"},
    "懐石": {"zh": "会席", "ko": "가이세키", "en": "kaiseki"},
    "釜": {"zh": "釜", "ko": "가마솥", "en": "kama"},
    "鳥": {"zh": "鸡肉", "ko": "닭", "en": "chicken"},
    "ウナギ": {"zh": "鳗鱼", "ko": "장어", "en": "eel"},
    "うな": {"zh": "鳗鱼", "ko": "장어", "en": "eel"},
    "そば": {"zh": "荞麦面", "ko": "소바", "en": "soba"},
    "うどん": {"zh": "乌冬", "ko": "우동", "en": "udon"},
    "温": {"zh": "热", "ko": "온", "en": "hot"},
    "冷": {"zh": "冷", "ko": "냉", "en": "cold"},
    "ミニ": {"zh": "迷你", "ko": "미니", "en": "mini"},
    "変更": {"zh": "变更", "ko": "변경", "en": "change"},
    "抹茶": {"zh": "抹茶", "ko": "말차", "en": "matcha"},
    "アイス": {"zh": "冰淇淋", "ko": "아이스크림", "en": "ice cream"},
}


def load_menu_translation_overrides() -> dict[str, dict[str, str]]:
    """
    menu_translations.txt の上書き翻訳を読む。
    形式: 日本語|中文|한국어|English
    """
    overrides: dict[str, dict[str, str]] = {}
    if not MENU_TRANSLATIONS_TXT_PATH.exists():
        return overrides
    for line in MENU_TRANSLATIONS_TXT_PATH.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        parts = [p.strip() for p in s.split("|")]
        if len(parts) < 4 or not parts[0]:
            continue
        overrides[parts[0]] = {"ja": parts[0], "zh": parts[1], "ko": parts[2], "en": parts[3]}
    return overrides


def translate_menu_text(text: str, lang: str, overrides: dict[str, dict[str, str]] | None = None) -> str:
    src = str(text or "").strip()
    if lang == "ja" or not src:
        return src
    overrides = overrides or load_menu_translation_overrides()
    if src in overrides and overrides[src].get(lang):
        return overrides[src][lang]
    if src in TRANSLATION_EXACT and TRANSLATION_EXACT[src].get(lang):
        return TRANSLATION_EXACT[src][lang]

    result = src
    for key in sorted(TRANSLATION_TERMS.keys(), key=len, reverse=True):
        val = TRANSLATION_TERMS[key].get(lang)
        if val:
            result = result.replace(key, val)
    if result == src:
        return src
    if lang == "en":
        result = re.sub(r"\s+", " ", result).strip()
        # 日本語由来の記号を少し整える
        result = result.replace("＋", " + ").replace("+", " + ")
    return result


def i18n_text(text: str, overrides: dict[str, dict[str, str]] | None = None) -> dict[str, str]:
    overrides = overrides or load_menu_translation_overrides()
    src = str(text or "")
    return {
        "ja": src,
        "zh": translate_menu_text(src, "zh", overrides),
        "ko": translate_menu_text(src, "ko", overrides),
        "en": translate_menu_text(src, "en", overrides),
    }


def add_i18n_to_products(products: list[dict]) -> list[dict]:
    overrides = load_menu_translation_overrides()
    result: list[dict] = []
    for p in products:
        item = dict(p)
        item["name_i18n"] = i18n_text(item.get("name", ""), overrides)
        item["category_i18n"] = i18n_text(item.get("category", ""), overrides)
        groups = []
        for group in item.get("option_groups", []) or []:
            g = dict(group)
            opts = []
            for opt in g.get("options", []) or []:
                o = dict(opt)
                o["name_i18n"] = i18n_text(o.get("name", ""), overrides)
                opts.append(o)
            g["options"] = opts
            groups.append(g)
        item["option_groups"] = groups
        result.append(item)
    return result


def build_category_view(products: list[dict]) -> list[dict]:
    categories = []
    seen = set()
    for p in products:
        category = p.get("category", "")
        if category not in seen:
            categories.append({"name": category, "i18n": p.get("category_i18n") or i18n_text(category)})
            seen.add(category)
    return categories


def order_name_i18n(product_name: str) -> dict[str, str]:
    """注文履歴の商品名（親商品 / 選択肢）を翻訳表示する。"""
    overrides = load_menu_translation_overrides()
    parts = [p.strip() for p in str(product_name or "").split(" / ") if p.strip()]
    if not parts:
        return i18n_text(product_name, overrides)
    return {
        lang: " / ".join(translate_menu_text(part, lang, overrides) if lang != "ja" else part for part in parts)
        for lang in ("ja", "zh", "ko", "en")
    }




def get_lan_ip() -> str:
    """
    同一Wi-Fi内のスマホからアクセスするためのPC側LAN IPを推定する。
    失敗時は127.0.0.1に戻す。
    """
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def get_base_url() -> str:
    """QRコードやスマホ用URLに使うベースURL。config.txt の BASE_URL があれば優先する。"""
    config = load_config()
    base_url = config.get("BASE_URL", "").strip()
    if base_url:
        return base_url.rstrip("/")
    return f"http://{get_lan_ip()}:{APP_PORT}"


HAGAKI_SIZE = (100 * mm, 148 * mm)


def load_or_create_secret_key() -> str:
    """Flaskセッション用の秘密鍵。config.txt の SECRET_KEY を優先し、未設定時はローカルファイルへ自動生成する。"""
    env_value = os.environ.get("KAKASHI_SECRET_KEY", "").strip()
    if env_value:
        return env_value

    config_value = load_config().get("SECRET_KEY", "").strip()
    if config_value:
        return config_value

    secret_file = BASE_DIR / ".flask_secret_key"
    try:
        if secret_file.exists():
            saved = secret_file.read_text(encoding="utf-8").strip()
            if saved:
                return saved
        generated = secrets.token_hex(32)
        secret_file.write_text(generated, encoding="utf-8")
        return generated
    except OSError:
        return secrets.token_hex(32)


app = Flask(__name__)
app.secret_key = load_or_create_secret_key()
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=str(load_config().get("SESSION_COOKIE_SECURE", "0")).strip().lower() in {"1", "true", "yes", "on"},
)


def db() -> sqlite3.Connection:
    # 21卓＋複数店員の同時操作を想定し、SQLiteのロック待機時間を設定する。
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 30000")
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def init_font() -> str:
    # ReportLabのCIDフォントを使うと、配布環境に日本語TTFを同梱しなくても
    # PDF内の日本語が文字化けしにくい。
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
        return "HeiseiKakuGo-W5"
    except Exception:
        pass

    candidates = [
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/ipafont-gothic/ipag.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for font_path in candidates:
        if Path(font_path).exists():
            try:
                pdfmetrics.registerFont(TTFont("JPFont", font_path))
                return "JPFont"
            except Exception:
                continue
    return "Helvetica"


PDF_FONT = init_font()





def ensure_column(conn: sqlite3.Connection, table_name: str, column_name: str, column_def: str) -> None:
    """既存DB向けの簡易マイグレーション。列がなければ追加する。"""
    columns = [row["name"] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()]
    if column_name not in columns:
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_def}")


def table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone()
    return row is not None


def migrate_existing_db(conn: sqlite3.Connection) -> None:
    """
    旧バージョンDBを壊さず、現行アプリで必要な列を補完する。
    SQLiteは列追加はできるが、既存列の型変更や制約変更は限定的なので、
    本試作では不足列の追加に絞る。
    """
    # tables
    if table_exists(conn, "tables"):
        ensure_column(conn, "tables", "status", "TEXT NOT NULL DEFAULT 'available'")
        ensure_column(conn, "tables", "current_receipt_no", "TEXT")

    # receipts
    if table_exists(conn, "receipts"):
        ensure_column(conn, "receipts", "table_no", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "token", "TEXT")
        ensure_column(conn, "receipts", "qr_path", "TEXT")
        ensure_column(conn, "receipts", "pdf_path", "TEXT")
        ensure_column(conn, "receipts", "checkout_pdf_path", "TEXT")
        ensure_column(conn, "receipts", "status", "TEXT NOT NULL DEFAULT 'active'")
        ensure_column(conn, "receipts", "subtotal", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "discount_amount", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "coupon_amount", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "discount_rate", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "total_amount", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "checkout_type", "TEXT")
        ensure_column(conn, "receipts", "payment_method", "TEXT")
        ensure_column(conn, "receipts", "manager_no", "TEXT")
        ensure_column(conn, "receipts", "created_at", "TEXT")
        ensure_column(conn, "receipts", "paid_at", "TEXT")
        ensure_column(conn, "receipts", "voided_at", "TEXT")
        ensure_column(conn, "receipts", "adult_count", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "child_count", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "child_chopsticks", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "spoon_fork", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "guest_info_submitted", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "guest_info_pdf_path", "TEXT")
        ensure_column(conn, "receipts", "adult_count", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "child_count", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "child_chopsticks", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "spoon_fork", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "guest_info_submitted", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "refunded_at", "TEXT")
        ensure_column(conn, "receipts", "refund_reason", "TEXT")
        ensure_column(conn, "receipts", "locked_by", "TEXT")
        ensure_column(conn, "receipts", "locked_at", "TEXT")
        conn.execute("UPDATE receipts SET status = 'active' WHERE status IS NULL OR status = ''")
        conn.execute("UPDATE receipts SET created_at = ? WHERE created_at IS NULL OR created_at = ''", (now(),))

    # products
    if table_exists(conn, "products"):
        ensure_column(conn, "products", "name", "TEXT")
        ensure_column(conn, "products", "category", "TEXT NOT NULL DEFAULT '料理'")
        ensure_column(conn, "products", "price", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "products", "is_active", "INTEGER NOT NULL DEFAULT 1")
        conn.execute("UPDATE products SET category = '料理' WHERE category IS NULL OR category = ''")
        conn.execute("UPDATE products SET is_active = 1 WHERE is_active IS NULL")
        conn.execute("UPDATE products SET price = 0 WHERE price IS NULL")

    # orders
    if table_exists(conn, "orders"):
        ensure_column(conn, "orders", "receipt_no", "TEXT")
        ensure_column(conn, "orders", "table_no", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "orders", "status", "TEXT NOT NULL DEFAULT 'ordered'")
        ensure_column(conn, "orders", "total_amount", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "orders", "created_at", "TEXT")
        conn.execute("UPDATE orders SET status = 'ordered' WHERE status IS NULL OR status = ''")
        conn.execute("UPDATE orders SET total_amount = 0 WHERE total_amount IS NULL")
        conn.execute("UPDATE orders SET created_at = ? WHERE created_at IS NULL OR created_at = ''", (now(),))

    # order_items
    if table_exists(conn, "order_items"):
        ensure_column(conn, "order_items", "receipt_no", "TEXT")
        ensure_column(conn, "order_items", "created_at", "TEXT")
        ensure_column(conn, "order_items", "order_id", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "order_items", "product_id", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "order_items", "product_name", "TEXT")
        ensure_column(conn, "order_items", "unit_price", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "order_items", "quantity", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "order_items", "line_total", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "order_items", "paid_quantity", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "order_items", "paid_at", "TEXT")
        conn.execute("UPDATE order_items SET product_name = '' WHERE product_name IS NULL")
        conn.execute("UPDATE order_items SET unit_price = 0 WHERE unit_price IS NULL")
        conn.execute("UPDATE order_items SET quantity = 0 WHERE quantity IS NULL")
        conn.execute("UPDATE order_items SET line_total = unit_price * quantity WHERE line_total IS NULL OR line_total = 0")

    # order_ticket_logs
    if table_exists(conn, "order_ticket_logs"):
        ensure_column(conn, "order_ticket_logs", "receipt_no", "TEXT")
        ensure_column(conn, "order_ticket_logs", "table_no", "INTEGER")
        ensure_column(conn, "order_ticket_logs", "issue_no", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "order_ticket_logs", "display_time", "TEXT")
        ensure_column(conn, "order_ticket_logs", "items_json", "TEXT")
        ensure_column(conn, "order_ticket_logs", "pdf_path", "TEXT")
        ensure_column(conn, "order_ticket_logs", "created_at", "TEXT")

    # register_payments
    if table_exists(conn, "register_payments"):
        ensure_column(conn, "register_payments", "register_log_id", "INTEGER")
        ensure_column(conn, "register_payments", "receipt_no", "TEXT")
        ensure_column(conn, "register_payments", "payment_method", "TEXT")
        ensure_column(conn, "register_payments", "amount", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "register_payments", "created_at", "TEXT")

    # register_logs
    if table_exists(conn, "register_logs"):
        ensure_column(conn, "register_logs", "receipt_no", "TEXT")
        ensure_column(conn, "register_logs", "manager_no", "TEXT")
        ensure_column(conn, "register_logs", "checkout_type", "TEXT")
        ensure_column(conn, "register_logs", "payment_method", "TEXT")
        ensure_column(conn, "register_logs", "subtotal", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "register_logs", "discount_rate", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "register_logs", "discount_amount", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "register_logs", "coupon_amount", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "register_logs", "total_amount", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "register_logs", "created_at", "TEXT")




def seed_sample_products(conn: sqlite3.Connection) -> None:
    """ブラックボックス試験用の商品を不足時に補完する。"""
    sample_products = [
        ("焼き鳥盛り合わせ", "料理", 780),
        ("唐揚げ", "料理", 650),
        ("枝豆", "料理", 320),
        ("ポテトフライ", "料理", 480),
        ("刺身三点盛り", "料理", 980),
        ("焼きおにぎり", "料理", 280),
        ("生ビール", "ドリンク", 550),
        ("ハイボール", "ドリンク", 480),
        ("レモンサワー", "ドリンク", 450),
        ("ウーロン茶", "ドリンク", 300),
        ("コーラ", "ドリンク", 300),
        ("バニラアイス", "デザート", 350),
    ]

    for name, category, price in sample_products:
        exists = conn.execute("SELECT id FROM products WHERE name = ?", (name,)).fetchone()
        if exists is None:
            conn.execute(
                "INSERT INTO products(name, category, price, is_active) VALUES (?, ?, ?, 1)",
                (name, category, price),
            )





def sync_product_menu_to_products(conn: sqlite3.Connection) -> None:
    """
    product_menu.txt の商品IDを products.id に同期する。

    order_items.product_id には外部キー制約がある環境があるため、
    TXTメニューのIDだけで注文をINSERTすると FOREIGN KEY constraint failed になる。
    注文保存前にTXT商品を products テーブルへ同じIDで同期し、
    本番注文・試験用メニューの両方で外部キー違反を防ぐ。
    """
    if not table_exists(conn, "products"):
        return

    ensure_column(conn, "products", "category", "TEXT NOT NULL DEFAULT '料理'")
    ensure_column(conn, "products", "is_active", "INTEGER NOT NULL DEFAULT 1")
    ensure_column(conn, "products", "price", "INTEGER NOT NULL DEFAULT 0")

    for product in load_product_menu(include_inactive=True):
        product_id = int(product["id"])
        is_active = 0 if product.get("inactive") else 1
        try:
            conn.execute(
                """
                INSERT INTO products(id, name, category, price, is_active)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(id) DO UPDATE SET
                    name = excluded.name,
                    category = excluded.category,
                    price = excluded.price,
                    is_active = excluded.is_active
                """,
                (product_id, product["name"], product["category"], int(product["base_price"]), is_active),
            )
        except sqlite3.OperationalError:
            # 古いSQLite互換。UPSERTが使えない環境でも同期できるようにする。
            conn.execute(
                "INSERT OR IGNORE INTO products(id, name, category, price, is_active) VALUES (?, ?, ?, ?, ?)",
                (product_id, product["name"], product["category"], int(product["base_price"]), is_active),
            )
            conn.execute(
                """
                UPDATE products
                SET name = ?, category = ?, price = ?, is_active = ?
                WHERE id = ?
                """,
                (product["name"], product["category"], int(product["base_price"]), is_active, product_id),
            )


def table_column_names(conn: sqlite3.Connection, table_name: str) -> set[str]:
    return {str(row["name"]) for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()}


def ensure_orders_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            receipt_no TEXT NOT NULL,
            table_no INTEGER NOT NULL,
            status TEXT NOT NULL DEFAULT 'ordered',
            total_amount INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            FOREIGN KEY(receipt_no) REFERENCES receipts(receipt_no)
        )
        """
    )


def insert_built_order_items(conn: sqlite3.Connection, receipt: sqlite3.Row, built_items: list[dict]) -> list[dict]:
    """
    注文行を現在のDBスキーマに合わせて安全にINSERTする。

    旧DBでは order_items が receipt_no 中心、別版DBでは order_id 外部キー付き、
    という差があるため、列の有無を見てINSERT文を切り替える。
    """
    if not built_items:
        return []

    sync_product_menu_to_products(conn)
    columns = table_column_names(conn, "order_items")
    has_order_id = "order_id" in columns

    order_id = None
    if has_order_id:
        ensure_orders_table(conn)
        total_amount = sum(int(item.get("line_total", 0) or 0) for item in built_items)
        cur = conn.execute(
            """
            INSERT INTO orders(receipt_no, table_no, status, total_amount, created_at)
            VALUES (?, ?, 'ordered', ?, ?)
            """,
            (receipt["receipt_no"], int(receipt["table_no"]), total_amount, now()),
        )
        order_id = int(cur.lastrowid)

    submitted_items = []
    for built in built_items:
        product_id = int(built["product_id"])
        if has_order_id:
            conn.execute(
                """
                INSERT INTO order_items(
                    order_id, receipt_no, product_id, product_name, unit_price,
                    quantity, line_total, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    order_id, receipt["receipt_no"], product_id, built["product_name"],
                    int(built["unit_price"]), int(built["quantity"]),
                    int(built["line_total"]), now(),
                ),
            )
        else:
            conn.execute(
                """
                INSERT INTO order_items(
                    receipt_no, product_id, product_name, unit_price,
                    quantity, line_total, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    receipt["receipt_no"], product_id, built["product_name"],
                    int(built["unit_price"]), int(built["quantity"]),
                    int(built["line_total"]), now(),
                ),
            )

        submitted_items.append({
            "product_name": built["product_name"],
            "quantity": int(built["quantity"]),
        })

    return submitted_items





class ProductMenuError(Exception):
    pass


def normalize_price_text(value: str) -> int:
    s = str(value).replace(",", "").replace("円", "").replace("¥", "").replace("￥", "").replace("(", "").replace(")", "").strip()
    if not s:
        raise ProductMenuError("金額が空です")
    sign = -1 if s.startswith("-") else 1
    s = s.lstrip("+-")
    digits = "".join(ch for ch in s if ch.isdigit())
    if not digits:
        raise ProductMenuError(f"金額を読み取れません: {value}")
    return sign * int(digits)


def split_top_level_commas(text: str) -> list[str]:
    parts = []
    buf = []
    depth = 0
    for ch in text:
        if ch == "{":
            depth += 1; buf.append(ch)
        elif ch == "}":
            depth -= 1
            if depth < 0:
                raise ProductMenuError("波括弧の閉じ方が不正です")
            buf.append(ch)
        elif ch == "," and depth == 0:
            parts.append("".join(buf).strip()); buf = []
        else:
            buf.append(ch)
    if depth != 0:
        raise ProductMenuError("波括弧の数が一致していません")
    if buf: parts.append("".join(buf).strip())
    return parts


def parse_option_text(option_text: str) -> dict:
    s = option_text.strip()
    if not s:
        raise ProductMenuError("空の選択肢があります")
    m = re.match(r"^(.*?)\(([-+]?[\d,]+円)\)$", s)
    if not m:
        raise ProductMenuError(f"選択肢の金額がありません: {s}")
    name = m.group(1).strip()
    if not name:
        raise ProductMenuError(f"選択肢名がありません: {s}")
    return {"name": name, "price_delta": normalize_price_text(m.group(2))}


def normalize_menu_image_path(value: str) -> str:
    """
    商品マスタの画像指定を正規化する。

    現行の標準:
      |001.png                 -> /photos/001.png
      |画像=001.png            -> /photos/001.png
      |photos/001.png          -> /photos/001.png
      |/photos/001.png         -> /photos/001.png

    互換維持:
      |static/menu_photos/001.png
      |menu_photos/001.png
      |static/menu/001.png
      |menu/001.png
    """
    s = str(value or "").strip().replace("\\", "/")
    if not s:
        return ""

    # 画像= / image= / img= が付いていても、付いていなくても受け付ける。
    for prefix in ("画像=", "image=", "img="):
        if s.startswith(prefix):
            s = s.split("=", 1)[1].strip().replace("\\", "/")
            break

    if not s:
        return ""

    lower = s.lower().split("?", 1)[0]
    image_exts = (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".svg")
    if not lower.endswith(image_exts):
        return ""

    # URL / 絶対URLはそのまま使う。
    if s.startswith(("http://", "https://")):
        return s

    # プロジェクト直下 photos フォルダを標準にする。
    # 例: Y:/.../Ver43/photos/001.png -> /photos/001.png
    if s.startswith("/photos/"):
        return s
    if s.startswith("photos/"):
        return "/" + s

    # static配下指定は互換維持する。
    if s.startswith(("/static/", "static/")):
        return s

    # 以前の menu_photos / menu_photo 指定も互換維持する。
    if s.startswith("menu_photos/"):
        return "static/" + s
    if s.startswith("menu_photo/"):
        return "static/menu_photos/" + s.split("/", 1)[1]

    # 旧指定 menu/xxx は互換維持する。
    if s.startswith("menu/"):
        return "static/" + s

    # ファイル名だけなら、プロジェクト直下 photos 配下として扱う。
    return "/photos/" + s


def apply_product_menu_attr(attr: str, inactive: bool, image_path: str) -> tuple[bool, str]:
    """商品マスタの | 属性を inactive / image_path に反映する。"""
    s = str(attr or "").strip()
    if not s:
        return inactive, image_path

    if s in {"停止", "販売停止", "売切れ", "売り切れ"}:
        return True, image_path

    normalized_image = normalize_menu_image_path(s)
    if normalized_image:
        return inactive, normalized_image

    return inactive, image_path


def parse_product_menu_line(line: str, menu_id: int, line_no: int | None = None) -> dict | None:
    """
    商品TXT 1行を辞書化する。

    基本:
      {[ジャンル]商品名(価格円)}

    画像:
      {[ジャンル]商品名(価格円)|画像=static/menu/sample.jpg}
      {[ジャンル]商品名(価格円)}|画像=static/menu/sample.jpg
      {[ジャンル]商品名(価格円)}|sample.jpg

    販売停止・売切れ:
      {[ジャンル]商品名(価格円)|停止}
      {[ジャンル]商品名(価格円)}|売切れ
    """
    raw = line.strip()
    if not raw or raw.startswith("#"):
        return None

    inactive = False
    image_path = ""

    # 外側 } の後ろに |画像=... |sample.jpg |停止 が付く形式を許可する
    suffix_parts = []
    if raw.startswith("{") and "}" in raw:
        close_idx = raw.rfind("}")
        suffix = raw[close_idx + 1:].strip()
        if suffix.startswith("|"):
            suffix_parts.extend([p.strip() for p in suffix.split("|") if p.strip()])
            raw = raw[:close_idx + 1].strip()

    for attr in suffix_parts:
        inactive, image_path = apply_product_menu_attr(attr, inactive, image_path)

    if not (raw.startswith("{") and raw.endswith("}")):
        raise ProductMenuError("行全体を { ... } で囲んでください")

    inner = raw[1:-1].strip()
    parts = split_top_level_commas(inner)
    if not parts:
        raise ProductMenuError("商品定義が空です")

    head = parts[0].strip()

    # 商品ヘッダー内の |画像=... |sample.jpg |停止 も許可する
    if "|" in head:
        head_parts = [p.strip() for p in head.split("|")]
        head = head_parts[0]
        for attr in head_parts[1:]:
            inactive, image_path = apply_product_menu_attr(attr, inactive, image_path)

    m = re.match(r"^\[(.*?)\](.*?)\(([\d,]+円)\)$", head)
    if not m:
        raise ProductMenuError("商品は [ジャンル]商品名(価格円) の形式で書いてください。価格は必須です")

    category = m.group(1).strip()
    name = m.group(2).strip()

    if not category:
        raise ProductMenuError("ジャンルが空です")
    if not name:
        raise ProductMenuError("商品名が空です")

    base_price = normalize_price_text(m.group(3))

    groups = []
    for group_index, group_text in enumerate(parts[1:]):
        g = group_text.strip()
        if not (g.startswith("{") and g.endswith("}")):
            raise ProductMenuError("選択肢グループは { ... } で囲んでください")

        g_inner = g[1:-1].strip()
        if not g_inner:
            raise ProductMenuError("空の選択肢グループがあります")

        option_parts = split_top_level_commas(g_inner)
        options = [parse_option_text(x) for x in option_parts if x.strip()]

        if not options:
            raise ProductMenuError("選択肢がありません")

        groups.append({"group_index": group_index, "required": True, "options": options})

    return {
        "id": menu_id,
        "category": category,
        "name": name,
        "base_price": base_price,
        "option_groups": groups,
        "inactive": inactive,
        "line_no": line_no,
        "image_path": image_path,
    }


def validate_product_menu() -> list[str]:
    errors=[]
    if not MENU_TXT_PATH.exists(): return [f"{MENU_TXT_PATH.name} が存在しません"]
    active_count=0; menu_id=1
    for line_no, line in enumerate(MENU_TXT_PATH.read_text(encoding="utf-8").splitlines(), start=1):
        s=line.strip()
        if not s or s.startswith("#"): continue
        try:
            item=parse_product_menu_line(s, menu_id, line_no)
            if item:
                menu_id += 1
                if not item.get("inactive"): active_count += 1
        except ProductMenuError as e:
            errors.append(f"{MENU_TXT_PATH.name} {line_no}行目: {e}")
    if active_count == 0:
        errors.append(f"{MENU_TXT_PATH.name}: 有効な商品が1件もありません")
    return errors


def print_product_menu_validation() -> None:
    errors=validate_product_menu()
    if errors:
        print("="*60); print("product_menu.txt にエラーがあります。修正してください。")
        for error in errors: print(error)
        print("="*60)
        raise RuntimeError("product_menu.txt の書式エラー")


def load_product_menu(include_inactive: bool = False) -> list[dict]:
    errors=validate_product_menu()
    if errors: raise RuntimeError("\n".join(errors))
    products=[]; menu_id=1
    for line_no, line in enumerate(MENU_TXT_PATH.read_text(encoding="utf-8").splitlines(), start=1):
        item=parse_product_menu_line(line, menu_id, line_no)
        if item:
            menu_id += 1
            if include_inactive or not item.get("inactive"):
                products.append(item)
    return products


def get_product_from_menu(menu_id: int) -> dict | None:
    for p in load_product_menu(include_inactive=False):
        if int(p["id"]) == int(menu_id): return p
    return None


def build_order_item_from_cart(cart_item: dict) -> dict | None:
    try:
        menu_id=int(cart_item.get("product_id")); qty=int(cart_item.get("quantity",0))
    except (TypeError, ValueError):
        return None
    if qty <= 0: return None
    # 数量は勝手に20へ丸めない。0以下のみ拒否する。
    product=get_product_from_menu(menu_id)
    if not product: return None
    selected_options=cart_item.get("selected_options", [])
    if not isinstance(selected_options, list): selected_options=[]
    if len(selected_options) != len(product["option_groups"]): return None
    option_names=[]; unit_price=int(product["base_price"])
    for group, selected_index in zip(product["option_groups"], selected_options):
        try: idx=int(selected_index)
        except (TypeError, ValueError): return None
        if idx < 0 or idx >= len(group["options"]): return None
        option=group["options"][idx]
        option_names.append(option["name"]); unit_price += int(option["price_delta"])
    display_name=product["name"]
    if option_names: display_name += " / " + " / ".join(option_names)
    return {"product_id": menu_id, "product_name": display_name, "unit_price": unit_price, "quantity": qty, "line_total": unit_price*qty}





def get_product_from_menu_any(menu_id: int) -> dict | None:
    for p in load_product_menu(include_inactive=True):
        if int(p["id"]) == int(menu_id):
            return p
    return None


def build_order_item_from_cart_detailed(cart_item: dict) -> tuple[dict | None, str | None]:
    """
    カート1行を検証し、失敗理由も返す。
    販売停止・売切れ後に客が送信した場合、分かりやすいメッセージを出すために使う。
    """
    try:
        menu_id = int(cart_item.get("product_id"))
        qty = int(cart_item.get("quantity", 0))
    except (TypeError, ValueError):
        return None, "不正な商品データが含まれていました。画面を更新して再度選択してください。"

    if qty <= 0:
        return None, "数量が0の商品が含まれていました。"
    # 数量は勝手に20へ丸めない。0以下のみ拒否する。

    product = get_product_from_menu_any(menu_id)
    if not product:
        return None, "商品マスタに存在しない商品が含まれていました。画面を更新して再度選択してください。"

    if product.get("inactive"):
        return None, f"販売が中止になった商品が含まれていました: {product['name']}。画面を更新して選び直してください。"

    selected_options = cart_item.get("selected_options", [])
    if not isinstance(selected_options, list):
        selected_options = []

    if len(selected_options) != len(product["option_groups"]):
        return None, f"{product['name']} の必須選択項目が不足しています。"

    option_names = []
    unit_price = int(product["base_price"])

    for group, selected_index in zip(product["option_groups"], selected_options):
        try:
            idx = int(selected_index)
        except (TypeError, ValueError):
            return None, f"{product['name']} の選択項目が不正です。"
        if idx < 0 or idx >= len(group["options"]):
            return None, f"{product['name']} の選択項目が不正です。"
        option = group["options"][idx]
        option_names.append(option["name"])
        unit_price += int(option["price_delta"])

    display_name = product["name"]
    if option_names:
        display_name += " / " + " / ".join(option_names)

    return {
        "product_id": menu_id,
        "product_name": display_name,
        "unit_price": unit_price,
        "quantity": qty,
        "line_total": unit_price * qty,
    }, None


def build_order_items_from_cart_with_errors(cart: list) -> tuple[list[dict], list[str]]:
    built_items = []
    errors = []
    for cart_item in cart:
        built, error = build_order_item_from_cart_detailed(cart_item)
        if built:
            built_items.append(built)
        elif error:
            errors.append(error)
    return built_items, errors


def find_latest_order_ticket_pdf(receipt_no: str) -> Path | None:
    if not PDF_DIR.exists():
        return None
    files = [p for p in PDF_DIR.glob(f"*_オーダー_{receipt_no}.pdf") if p.is_file()]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)



def format_order_ticket_display_name(product_name: str) -> list[str]:
    """
    オーダー票表示用。
    親商品と選択肢を「親商品」「　選択肢」の箇条書きに分解する。
    例:
    大エビ天丼面セット / 温うどん / なし
    -> ["大エビ天丼面セット", "　温うどん", "　なし"]
    """
    parts = [p.strip() for p in str(product_name or "").split(" / ") if p.strip()]
    if not parts:
        return [""]
    lines = [parts[0]]
    lines.extend(["　" + p for p in parts[1:]])
    return lines


def normalized_order_ticket_items(submitted_items: list[dict]) -> list[dict]:
    """
    DB保存・PDF再生成に使うため、オーダー票原本情報を正規化する。
    """
    result = []
    for item in submitted_items or []:
        try:
            qty = int(item.get("quantity", 0))
        except (TypeError, ValueError):
            qty = 0
        if qty <= 0:
            continue
        result.append({
            "product_name": str(item.get("product_name", "")),
            "quantity": qty,
        })
    return result


def save_order_ticket_log(
    receipt_no: str,
    table_no: int,
    issue_no: int,
    display_time: str,
    submitted_items: list[dict],
    pdf_path: str,
) -> None:
    """
    オーダー票の原本表示情報をDBへ保存する。
    再発行時はこの情報を使うため、PDFファイルが削除されてもNo/時刻/表示内容を保持できる。
    """
    items_json = json.dumps(normalized_order_ticket_items(submitted_items), ensure_ascii=False)
    rel_path = ""
    try:
        rel_path = str(Path(pdf_path).relative_to(BASE_DIR)).replace("\\", "/")
    except Exception:
        rel_path = str(pdf_path)

    with db() as conn:
        migrate_existing_db(conn)
        conn.execute(
            """
            INSERT INTO order_ticket_logs(
                receipt_no, table_no, issue_no, display_time,
                items_json, pdf_path, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (receipt_no, int(table_no), int(issue_no), display_time, items_json, rel_path, now()),
        )


def latest_order_ticket_log(receipt_no: str):
    """
    指定伝票番号の最新オーダー票原本情報を返す。
    伝票番号だけで再発行する仕様のため、複数回追加注文がある場合は最新オーダー票を対象にする。
    """
    with db() as conn:
        migrate_existing_db(conn)
        return conn.execute(
            """
            SELECT *
            FROM order_ticket_logs
            WHERE receipt_no = ?
            ORDER BY created_at DESC, id DESC
            LIMIT 1
            """,
            (receipt_no,),
        ).fetchone()


def render_order_ticket_pdf(
    receipt_no: str,
    table_no: int,
    submitted_items: list[dict],
    issue_no: int,
    display_time: str,
    filename_prefix: str = "",
) -> str:
    """
    オーダー票PDFを描画する共通処理。
    issue_no/display_timeを引数で受け取るため、再発行でも原本と同じ表示にできる。
    """
    items = normalized_order_ticket_items(submitted_items)
    if not items:
        return ""

    PDF_DIR.mkdir(parents=True, exist_ok=True)
    DAILY_REPORT_DIR.mkdir(parents=True, exist_ok=True)

    base_name = order_ticket_filename(receipt_no)
    if filename_prefix:
        pdf_file = PDF_DIR / f"{filename_prefix}_{base_name}"
    else:
        pdf_file = PDF_DIR / base_name

    # 商品表示は親商品＋選択肢の複数行になるため、表示行数で高さを計算する。
    display_rows = 0
    for item in items:
        display_rows += len(format_order_ticket_display_name(item["product_name"]))

    width = 80 * mm
    height = max(62 * mm, (30 + 6 * display_rows + 10 * len(items) + 14) * mm)

    c = canvas.Canvas(str(pdf_file), pagesize=(width, height))
    w, h = width, height

    margin_x = 8 * mm
    y = h - 9 * mm

    c.setFont(PDF_FONT, 13)
    c.drawCentredString(w / 2, y, "オーダー票")

    y -= 9 * mm
    c.setFont(PDF_FONT, 9)
    c.drawString(margin_x, y, f"No.{format_issue_no(issue_no)}")
    c.drawRightString(w - margin_x, y, str(display_time))

    y -= 7 * mm
    c.setFont(PDF_FONT, 10)
    c.drawString(margin_x, y, f"テーブル番号 {table_no}")

    y -= 5 * mm
    c.line(margin_x, y, w - margin_x, y)
    y -= 8 * mm

    name_x = margin_x
    qty_x = w - margin_x
    name_max_width = 52 * mm

    for item in items:
        qty = int(item["quantity"])
        lines = format_order_ticket_display_name(item["product_name"])

        # 親商品行に数量を表示
        c.setFont(PDF_FONT, 10)
        draw_limited_text(c, lines[0], name_x, y, name_max_width, PDF_FONT, 10)
        c.setFont(PDF_FONT, 11)
        c.drawRightString(qty_x, y, str(qty))
        y -= 6 * mm

        # 選択肢行は1マス字下げ、数量は表示しない
        for opt in lines[1:]:
            c.setFont(PDF_FONT, 9)
            draw_limited_text(c, opt, name_x, y, name_max_width, PDF_FONT, 9)
            y -= 5.5 * mm

        y -= 2 * mm

    c.showPage()
    c.save()
    return str(pdf_file)


def reissue_order_ticket_pdf_original(receipt_no: str) -> str:
    """
    オーダー票再発行。
    PDFコピーではなく、DBに保存した原本情報から再生成する。
    No.・時刻・テーブル番号・商品表示・数量は原本と同じ値を使用する。
    """
    log = latest_order_ticket_log(receipt_no)
    if log is not None:
        try:
            items = json.loads(log["items_json"] or "[]")
        except json.JSONDecodeError:
            items = []

        if items:
            prefix = datetime.now().strftime("%Y_%m_%d_%H_%M_%S") + f"_{datetime.now().microsecond:06d}_再発行"
            return render_order_ticket_pdf(
                receipt_no=log["receipt_no"],
                table_no=int(log["table_no"]),
                submitted_items=items,
                issue_no=int(log["issue_no"]),
                display_time=str(log["display_time"]),
                filename_prefix=prefix,
            )

    # v35以前のDBで原本ログが無い場合の互換fallback。
    # 原本No/時刻はDBに無いため完全再現不可だが、注文内容だけは再生成する。
    table_no, submitted_items = build_submitted_items_from_receipt(receipt_no)
    if table_no is None or not submitted_items:
        return ""
    return generate_order_ticket_pdf(receipt_no=receipt_no, table_no=table_no, submitted_items=submitted_items)



def get_cleanup_days() -> int:
    config = load_config()
    try:
        days = int(config.get("CLEANUP_DAYS", "2") or 2)
    except ValueError:
        days = 2
    return max(1, days)


def cleanup_old_generated_files(days: int | None = None) -> None:
    """
    2日前以上のQR画像、QR付き入店伝票PDF、オーダー票PDF、会計PDFを削除する。
    DB本体は削除しない。過去会計データはDBに残る。
    """
    if days is None:
        days = get_cleanup_days()

    cutoff = time.time() - (days * 24 * 60 * 60)
    targets = []
    for folder, patterns in [(QR_DIR, ["*.png", "*.jpg", "*.jpeg"]), (PDF_DIR, ["*.pdf"])]:
        if not folder.exists():
            continue
        for pattern in patterns:
            targets.extend(folder.glob(pattern))

    deleted = 0
    for path in targets:
        try:
            # 会計レシートは永続保存対象。2日前削除から除外する。
            if path.suffix.lower() == ".pdf" and ("_会計_" in path.name or "_個別会計_" in path.name):
                continue
            if path.is_file() and path.stat().st_mtime < cutoff:
                path.unlink()
                deleted += 1
        except OSError:
            pass

    if deleted:
        print(f"古い生成ファイルを削除しました: {deleted}件 / {days}日前以上")


def list_receipt_generated_files(receipt_no: str) -> dict:
    files = {"entry_pdfs": [], "order_pdfs": [], "checkout_pdfs": [], "qr_images": []}

    if PDF_DIR.exists():
        for p in sorted(PDF_DIR.glob(f"*{receipt_no}*.pdf")):
            if "_オーダー_" in p.name:
                files["order_pdfs"].append(p)
            elif "_会計_" in p.name or "_個別会計_" in p.name:
                files["checkout_pdfs"].append(p)
            else:
                files["entry_pdfs"].append(p)

    if QR_DIR.exists():
        for p in sorted(QR_DIR.glob(f"*{receipt_no}*.*")):
            files["qr_images"].append(p)

    return files


def build_submitted_items_from_receipt(receipt_no: str) -> tuple[int | None, list[dict]]:
    with db() as conn:
        migrate_existing_db(conn)
        receipt = conn.execute("SELECT * FROM receipts WHERE receipt_no = ?", (receipt_no,)).fetchone()
        if receipt is None:
            return None, []

        rows = conn.execute(
            """
            SELECT product_name, SUM(quantity) AS quantity
            FROM order_items
            WHERE receipt_no = ?
            GROUP BY product_name
            ORDER BY MIN(id)
            """,
            (receipt_no,),
        ).fetchall()

    items = [
        {"product_name": row["product_name"], "quantity": int(row["quantity"] or 0)}
        for row in rows
        if int(row["quantity"] or 0) > 0
    ]
    return int(receipt["table_no"]), items


def init_db() -> None:
    print_product_menu_validation()
    cleanup_old_generated_files()
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    QR_DIR.mkdir(parents=True, exist_ok=True)
    PDF_DIR.mkdir(parents=True, exist_ok=True)

    with db() as conn:
        # WALにより、読み取りと書き込みの競合を減らす。
        conn.execute("PRAGMA journal_mode = WAL")
        conn.execute("PRAGMA synchronous = NORMAL")
        conn.execute("PRAGMA temp_store = MEMORY")

        ensure_auth_tables(conn)
        ensure_default_admin_user(conn)
        ensure_receipt_counter_table(conn)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS tables (
                table_no INTEGER PRIMARY KEY,
                status TEXT NOT NULL DEFAULT 'available',
                current_receipt_no TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS receipts (
                receipt_no TEXT PRIMARY KEY,
                table_no INTEGER NOT NULL,
                token TEXT NOT NULL UNIQUE,
                qr_path TEXT NOT NULL,
                pdf_path TEXT,
                checkout_pdf_path TEXT,
                status TEXT NOT NULL DEFAULT 'active',
                subtotal INTEGER NOT NULL DEFAULT 0,
                coupon_amount INTEGER NOT NULL DEFAULT 0,
                discount_rate INTEGER NOT NULL DEFAULT 0,
                discount_amount INTEGER NOT NULL DEFAULT 0,
                total_amount INTEGER NOT NULL DEFAULT 0,
                checkout_type TEXT,
                payment_method TEXT,
                manager_no TEXT,
                created_at TEXT NOT NULL,
                paid_at TEXT,
                voided_at TEXT,
                adult_count INTEGER NOT NULL DEFAULT 0,
                child_count INTEGER NOT NULL DEFAULT 0,
                child_chopsticks INTEGER NOT NULL DEFAULT 0,
                spoon_fork INTEGER NOT NULL DEFAULT 0,
                guest_info_submitted INTEGER NOT NULL DEFAULT 0,
                guest_info_pdf_path TEXT,
                FOREIGN KEY(table_no) REFERENCES tables(table_no)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                price INTEGER NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                receipt_no TEXT NOT NULL,
                product_id INTEGER NOT NULL,
                product_name TEXT NOT NULL,
                unit_price INTEGER NOT NULL,
                quantity INTEGER NOT NULL,
                line_total INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY(receipt_no) REFERENCES receipts(receipt_no),
                FOREIGN KEY(product_id) REFERENCES products(id)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS order_ticket_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                receipt_no TEXT NOT NULL,
                table_no INTEGER NOT NULL,
                issue_no INTEGER NOT NULL,
                display_time TEXT NOT NULL,
                items_json TEXT NOT NULL,
                pdf_path TEXT,
                created_at TEXT NOT NULL
            )
        """)

        ensure_order_cancel_tables(conn)
        ensure_daily_close_tables(conn)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS register_payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                register_log_id INTEGER,
                receipt_no TEXT NOT NULL,
                payment_method TEXT NOT NULL,
                amount INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            )
        """)


        conn.execute("""
            CREATE TABLE IF NOT EXISTS receipt_locks (
                receipt_no TEXT PRIMARY KEY,
                locked_by TEXT NOT NULL,
                locked_at TEXT NOT NULL,
                expires_at INTEGER NOT NULL
            )
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS cash_drawer_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                action_type TEXT NOT NULL,
                amount INTEGER NOT NULL,
                manager_no TEXT NOT NULL,
                note TEXT,
                created_at TEXT NOT NULL
            )
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS register_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                receipt_no TEXT NOT NULL,
                manager_no TEXT NOT NULL,
                checkout_type TEXT,
                payment_method TEXT NOT NULL,
                subtotal INTEGER NOT NULL,
                coupon_amount INTEGER NOT NULL DEFAULT 0,
                discount_rate INTEGER NOT NULL DEFAULT 0,
                discount_amount INTEGER NOT NULL DEFAULT 0,
                total_amount INTEGER NOT NULL,
                created_at TEXT NOT NULL
            )
        """)

        
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                category TEXT NOT NULL,
                price INTEGER NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                receipt_no TEXT NOT NULL,
                table_no INTEGER NOT NULL,
                status TEXT NOT NULL DEFAULT 'ordered',
                total_amount INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                FOREIGN KEY(receipt_no) REFERENCES receipts(receipt_no)
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER NOT NULL,
                product_id INTEGER NOT NULL,
                product_name TEXT NOT NULL,
                unit_price INTEGER NOT NULL,
                quantity INTEGER NOT NULL,
                line_total INTEGER NOT NULL,
                FOREIGN KEY(order_id) REFERENCES orders(id),
                FOREIGN KEY(product_id) REFERENCES products(id)
            )
            """
        )

        # 旧バージョンDB対策：
        # Ver7以前の試作DBに products テーブルが残っている場合、
        # category / is_active 列が存在せず起動エラーになるため自動追加する。
        ensure_column(conn, "products", "category", "TEXT NOT NULL DEFAULT '料理'")
        ensure_column(conn, "products", "is_active", "INTEGER NOT NULL DEFAULT 1")

        conn.execute("UPDATE products SET category = '料理' WHERE category IS NULL OR category = ''")
        conn.execute("UPDATE products SET is_active = 1 WHERE is_active IS NULL")


        conn.execute("""
            CREATE TABLE IF NOT EXISTS pdf_issue_counters (
                date_key TEXT NOT NULL,
                doc_type TEXT NOT NULL,
                last_no INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY(date_key, doc_type)
            )
        """)

        migrate_existing_db(conn)
        seed_sample_products(conn)
        sync_product_menu_to_products(conn)

        product_count = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        if product_count == 0:
            sample_products = [
                ("焼き鳥盛り合わせ", "料理", 780),
                ("唐揚げ", "料理", 650),
                ("枝豆", "料理", 320),
                ("ポテトフライ", "料理", 480),
                ("刺身三点盛り", "料理", 980),
                ("焼きおにぎり", "料理", 280),
                ("生ビール", "ドリンク", 550),
                ("ハイボール", "ドリンク", 480),
                ("レモンサワー", "ドリンク", 450),
                ("ウーロン茶", "ドリンク", 300),
                ("コーラ", "ドリンク", 300),
                ("バニラアイス", "デザート", 350),
            ]
            conn.executemany(
                "INSERT INTO products(name, category, price, is_active) VALUES (?, ?, ?, 1)",
                sample_products,
            )

        # 既存 receipts テーブルにも不足列があれば追加する。
        ensure_column(conn, "receipts", "subtotal", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "discount_amount", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "coupon_amount", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "discount_rate", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "total_amount", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "receipts", "checkout_type", "TEXT")
        ensure_column(conn, "receipts", "payment_method", "TEXT")
        ensure_column(conn, "receipts", "manager_no", "TEXT")
        ensure_column(conn, "receipts", "paid_at", "TEXT")
        ensure_column(conn, "receipts", "voided_at", "TEXT")
        ensure_column(conn, "receipts", "refunded_at", "TEXT")
        ensure_column(conn, "receipts", "refund_reason", "TEXT")
        ensure_column(conn, "receipts", "locked_by", "TEXT")
        ensure_column(conn, "receipts", "locked_at", "TEXT")
        ensure_column(conn, "receipts", "pdf_path", "TEXT")

        for i in range(1, TABLE_COUNT + 1):
            conn.execute(
                "INSERT OR IGNORE INTO tables(table_no, status) VALUES(?, 'available')",
                (i,),
            )

        # 旧バージョンDB対策：
        # Ver7以前の試作DBに products テーブルが残っている場合、
        # category / is_active 列が存在せず起動エラーになるため自動追加する。
        ensure_column(conn, "products", "category", "TEXT NOT NULL DEFAULT '料理'")
        ensure_column(conn, "products", "is_active", "INTEGER NOT NULL DEFAULT 1")

        conn.execute("UPDATE products SET category = '料理' WHERE category IS NULL OR category = ''")
        conn.execute("UPDATE products SET is_active = 1 WHERE is_active IS NULL")


        conn.execute("""
            CREATE TABLE IF NOT EXISTS pdf_issue_counters (
                date_key TEXT NOT NULL,
                doc_type TEXT NOT NULL,
                last_no INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY(date_key, doc_type)
            )
        """)

        migrate_existing_db(conn)
        seed_sample_products(conn)
        sync_product_menu_to_products(conn)

        product_count = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        if product_count == 0:
            products = [
                ("かかし定食", 980),
                ("唐揚げ定食", 880),
                ("焼き魚定食", 920),
                ("親子丼", 760),
                ("カレーライス", 700),
                ("うどん", 550),
                ("枝豆", 320),
                ("ポテト", 380),
                ("ソフトドリンク", 250),
                ("デザート", 420),
            ]
            conn.executemany("INSERT INTO products(name, price) VALUES(?, ?)", products)


def ensure_receipt_counter_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS receipt_counters (
            counter_name TEXT PRIMARY KEY,
            last_no INTEGER NOT NULL DEFAULT 0
        )
        """
    )


def current_max_receipt_no(conn: sqlite3.Connection) -> int:
    rows = conn.execute("SELECT receipt_no FROM receipts").fetchall()
    max_no = 0
    for row in rows:
        m = re.fullmatch(r"(\d{4})", str(row["receipt_no"]))
        if m:
            max_no = max(max_no, int(m.group(1)))
    return max_no


def next_receipt_no_in_conn(conn: sqlite3.Connection) -> str:
    """同時入店処理でも重複しないよう、単一トランザクション内で連番を進める。"""
    ensure_receipt_counter_table(conn)
    row = conn.execute(
        "SELECT last_no FROM receipt_counters WHERE counter_name = 'receipt_no'"
    ).fetchone()

    if row is None:
        next_no = current_max_receipt_no(conn) + 1
        conn.execute(
            "INSERT INTO receipt_counters(counter_name, last_no) VALUES('receipt_no', ?)",
            (next_no,),
        )
    else:
        next_no = max(int(row["last_no"] or 0), current_max_receipt_no(conn)) + 1
        conn.execute(
            "UPDATE receipt_counters SET last_no = ? WHERE counter_name = 'receipt_no'",
            (next_no,),
        )

    return f"{next_no:04d}"


def make_receipt_no(table_no: int) -> str:
    """0001から順番にレシート番号を発行する。複数端末同時押下でも重複しない。"""
    with db() as conn:
        conn.execute("BEGIN IMMEDIATE")
        return next_receipt_no_in_conn(conn)


def mobile_url(token: str) -> str:
    # 店内ローカル運用では、このホスト部分を店内サーバーIPに変更する。
    return f"{get_base_url()}/mobile/{token}"


def make_qr(receipt_no: str, token: str) -> str:
    img = qrcode.make(mobile_url(token))
    filename = f"{receipt_no}.png"
    img.save(QR_DIR / filename)
    return f"qr/{filename}"


def get_or_create_test_menu_receipt() -> sqlite3.Row:
    """スタッフ確認用。QRを読まずに開けるテーブル999固定の試験用伝票を返す。"""
    QR_DIR.mkdir(parents=True, exist_ok=True)

    # 既存DBマイグレーションは通常トランザクション外で済ませる。
    with db() as conn:
        migrate_existing_db(conn)

    with db() as conn:
        conn.execute("BEGIN IMMEDIATE")
        conn.execute(
            """
            INSERT OR IGNORE INTO tables(table_no, status, current_receipt_no)
            VALUES (?, 'occupied', NULL)
            """,
            (TEST_TABLE_NO,),
        )

        receipt = conn.execute(
            """
            SELECT *
            FROM receipts
            WHERE table_no = ? AND status = 'active'
            ORDER BY created_at DESC, receipt_no DESC
            LIMIT 1
            """,
            (TEST_TABLE_NO,),
        ).fetchone()

        if receipt is None:
            receipt_no = next_receipt_no_in_conn(conn)
            token = secrets.token_urlsafe(24)
            qr_path = make_qr(receipt_no, token)
            created_at = now()
            conn.execute(
                """
                INSERT INTO receipts(
                    receipt_no, table_no, token, qr_path, status, created_at,
                    adult_count, child_count, child_chopsticks, spoon_fork, guest_info_submitted
                ) VALUES (?, ?, ?, ?, 'active', ?, 1, 0, 0, 0, 1)
                """,
                (receipt_no, TEST_TABLE_NO, token, qr_path, created_at),
            )
            conn.execute(
                """
                UPDATE tables
                SET status = 'occupied', current_receipt_no = ?
                WHERE table_no = ?
                """,
                (receipt_no, TEST_TABLE_NO),
            )
            receipt = conn.execute("SELECT * FROM receipts WHERE receipt_no = ?", (receipt_no,)).fetchone()
        else:
            conn.execute(
                """
                UPDATE receipts
                SET guest_info_submitted = 1,
                    adult_count = CASE WHEN adult_count <= 0 THEN 1 ELSE adult_count END
                WHERE receipt_no = ?
                """,
                (receipt["receipt_no"],),
            )
            conn.execute(
                """
                UPDATE tables
                SET status = 'occupied', current_receipt_no = ?
                WHERE table_no = ?
                """,
                (receipt["receipt_no"], TEST_TABLE_NO),
            )
            receipt = conn.execute("SELECT * FROM receipts WHERE receipt_no = ?", (receipt["receipt_no"],)).fetchone()

        return receipt


def pdf_filename(receipt_no: str, created_at: str | None = None) -> str:
    """yyyy_MM_DD_レシート番号.pdf 形式で保存する。"""
    if created_at:
        try:
            dt = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            dt = datetime.now()
    else:
        dt = datetime.now()
    return f"{dt.strftime('%Y_%m_%d')}_{receipt_no}.pdf"


def generate_ticket_pdf(receipt_no: str) -> str:
    with db() as conn:
        receipt = conn.execute("SELECT * FROM receipts WHERE receipt_no = ?", (receipt_no,)).fetchone()
    if receipt is None:
        raise ValueError("receipt not found")

    pdf_file = PDF_DIR / pdf_filename(receipt_no, receipt["created_at"])
    qr_file = BASE_DIR / "static" / receipt["qr_path"]

    # はがきサイズ: 100mm x 148mm
    c = canvas.Canvas(str(pdf_file), pagesize=HAGAKI_SIZE)
    w, h = HAGAKI_SIZE

    margin_x = 12 * mm
    y = h - 14 * mm

    c.setFont(PDF_FONT, 16)
    c.drawCentredString(w / 2, y, STORE_NAME)

    y -= 16 * mm
    c.setFont(PDF_FONT, 11)
    c.drawCentredString(w / 2, y, "ご利用ありがとうございます")

    y -= 16 * mm
    c.setFont(PDF_FONT, 10.5)
    c.drawString(margin_x, y, "レシート番号")
    c.drawString(margin_x + 34 * mm, y, str(receipt["receipt_no"]))

    y -= 9 * mm
    c.drawString(margin_x, y, "テーブル番号")
    c.drawString(margin_x + 34 * mm, y, str(receipt["table_no"]))

    qr_size = 42 * mm
    qr_x = (w - qr_size) / 2
    qr_y = y - 50 * mm
    c.drawImage(ImageReader(str(qr_file)), qr_x, qr_y, width=qr_size, height=qr_size)

    y = qr_y - 9 * mm
    c.setFont(PDF_FONT, 9)
    c.drawCentredString(w / 2, y, "QRコードを読み取って注文してください。")

    y -= 9 * mm
    c.setFont(PDF_FONT, 8.5)
    c.drawCentredString(w / 2, y, f"発行日時: {receipt['created_at']}")

    c.showPage()
    c.save()

    rel = f"receipts/{pdf_file.name}"
    with db() as conn:
        conn.execute("UPDATE receipts SET pdf_path = ? WHERE receipt_no = ?", (rel, receipt_no))
    return str(pdf_file)



def checkout_pdf_filename(receipt_no: str, paid_at: str | None = None) -> str:
    """yyyy_MM_DD_会計_レシート番号.pdf 形式で保存する。"""
    if paid_at:
        try:
            dt = datetime.strptime(paid_at, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            dt = datetime.now()
    else:
        dt = datetime.now()
    return f"{dt.strftime('%Y_%m_%d')}_会計_{receipt_no}.pdf"


def draw_pdf_line(c: canvas.Canvas, text: str, x: float, y: float, max_width: float, font_name: str, font_size: float) -> None:
    """
    ReportLabの簡易1行描画。
    長すぎる商品名は末尾を省略して、金額欄との重なりを防ぐ。
    """
    c.setFont(font_name, font_size)
    s = str(text)
    while s and c.stringWidth(s, font_name, font_size) > max_width:
        s = s[:-1]
    if s != str(text):
        s = s[:-1] + "…"
    c.drawString(x, y, s)




def next_daily_issue_no(doc_type: str) -> int:
    """
    当日内のPDF発行番号を返す。
    date_key=YYYY-MM-DD で管理するため、毎日0時をまたぐと No.0001 から始まる。
    doc_typeごとに 0001, 0002... と増える。
    """
    date_key = datetime.now().strftime("%Y-%m-%d")
    with db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS pdf_issue_counters (
                date_key TEXT NOT NULL,
                doc_type TEXT NOT NULL,
                last_no INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY(date_key, doc_type)
            )
        """)
        row = conn.execute(
            "SELECT last_no FROM pdf_issue_counters WHERE date_key = ? AND doc_type = ?",
            (date_key, doc_type),
        ).fetchone()

        if row is None:
            issue_no = 1
            conn.execute(
                "INSERT INTO pdf_issue_counters(date_key, doc_type, last_no) VALUES (?, ?, ?)",
                (date_key, doc_type, issue_no),
            )
        else:
            issue_no = int(row["last_no"]) + 1
            conn.execute(
                "UPDATE pdf_issue_counters SET last_no = ? WHERE date_key = ? AND doc_type = ?",
                (issue_no, date_key, doc_type),
            )

    return issue_no


def format_issue_no(issue_no: int) -> str:
    return f"{issue_no:04d}"


def format_pdf_datetime(dt_text: str | None = None) -> str:
    """
    会計伝票用日時。
    ユーザー指定: YYYY:MM:DD_HH:MM:SS
    """
    if dt_text:
        try:
            dt = datetime.strptime(dt_text, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            dt = datetime.now()
    else:
        dt = datetime.now()
    return dt.strftime("%Y:%m:%d_%H:%M:%S")


def draw_limited_text(c: canvas.Canvas, text: str, x: float, y: float, max_width: float, font_name: str, font_size: float) -> None:
    """
    商品名が長すぎて個数・金額欄に重ならないように、描画幅を固定する。
    """
    c.setFont(font_name, font_size)
    s = str(text)
    if c.stringWidth(s, font_name, font_size) <= max_width:
        c.drawString(x, y, s)
        return

    ellipsis = "…"
    while s and c.stringWidth(s + ellipsis, font_name, font_size) > max_width:
        s = s[:-1]
    c.drawString(x, y, s + ellipsis)



def order_ticket_filename(receipt_no: str) -> str:
    dt = datetime.now()
    return f"{dt.strftime('%Y_%m_%d_%H_%M_%S')}_{dt.microsecond:06d}_オーダー_{receipt_no}.pdf"



def guest_info_pdf_filename(receipt_no: str) -> str:
    dt = datetime.now()
    return f"{dt.strftime('%Y_%m_%d_%H_%M_%S')}_{dt.microsecond:06d}_人数_{receipt_no}.pdf"


def generate_guest_info_pdf(
    receipt_no: str,
    table_no: int,
    adult_count: int,
    child_count: int,
    child_chopsticks: int,
    spoon_fork: int,
) -> str:
    """
    QR読み取り後の初回人数情報をPDF出力する。
    HH:MM + テーブル番号 + 人数 + 子供用備品を表示する。
    """
    PDF_DIR.mkdir(parents=True, exist_ok=True)

    pdf_file = PDF_DIR / guest_info_pdf_filename(receipt_no)

    width = 80 * mm
    height = 62 * mm

    c = canvas.Canvas(str(pdf_file), pagesize=(width, height))
    w, h = width, height
    margin_x = 8 * mm
    y = h - 9 * mm
    hh_mm = datetime.now().strftime("%H:%M")

    c.setFont(PDF_FONT, 13)
    c.drawCentredString(w / 2, y, "人数票")

    y -= 9 * mm
    c.setFont(PDF_FONT, 10)
    c.drawString(margin_x, y, f"{hh_mm}")
    c.drawRightString(w - margin_x, y, f"テーブル番号 {table_no}")

    y -= 5 * mm
    c.line(margin_x, y, w - margin_x, y)

    y -= 9 * mm
    c.setFont(PDF_FONT, 11)
    c.drawString(margin_x, y, f"大人 {int(adult_count)}名")

    y -= 8 * mm
    c.drawString(margin_x, y, f"子供 {int(child_count)}名")

    y -= 8 * mm
    c.drawString(margin_x, y, f"(子供用箸 {int(child_chopsticks)}個)")

    y -= 8 * mm
    c.drawString(margin_x, y, f"(子供用スプーン・フォーク {int(spoon_fork)}個)")

    c.showPage()
    c.save()

    rel_path = str(pdf_file.relative_to(BASE_DIR)).replace("\\", "/")
    with db() as conn:
        migrate_existing_db(conn)
        conn.execute(
            "UPDATE receipts SET guest_info_pdf_path = ? WHERE receipt_no = ?",
            (rel_path, receipt_no),
        )

    return str(pdf_file)



def generate_order_ticket_pdf(receipt_no: str, table_no: int, submitted_items: list[dict]) -> str:
    """
    厨房向けオーダー票PDFをバックエンドで自動保存する。
    新規注文・追加注文どちらも、今回送信された商品だけを表示する。
    原本情報は order_ticket_logs に保存し、再発行時に同じNo/時刻/表示内容を再現する。
    """
    items = normalized_order_ticket_items(submitted_items)
    if not items:
        return ""

    issue_no = next_daily_issue_no("order_ticket")
    display_time = datetime.now().strftime("%H:%M")

    pdf_path = render_order_ticket_pdf(
        receipt_no=receipt_no,
        table_no=table_no,
        submitted_items=items,
        issue_no=issue_no,
        display_time=display_time,
    )

    if pdf_path:
        save_order_ticket_log(
            receipt_no=receipt_no,
            table_no=table_no,
            issue_no=issue_no,
            display_time=display_time,
            submitted_items=items,
            pdf_path=pdf_path,
        )

    return pdf_path



def get_partial_receipt_display_no(receipt_no: str) -> str:
    """
    個別会計レシート表示用の番号を返す。
    register_logs に記録された個別会計回数を使い、0001-1, 0001-2 のように付番する。
    """
    with db() as conn:
        migrate_existing_db(conn)
        row = conn.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM register_logs
            WHERE receipt_no = ? AND checkout_type = '個別会計'
            """,
            (receipt_no,),
        ).fetchone()
    seq = int(row["cnt"] or 0)
    seq = max(1, seq)
    return f"{receipt_no}-{seq}"


def checkout_pdf_filename_custom(receipt_no: str, label: str = "会計") -> str:
    """
    会計伝票PDFの保存名。
    個別会計は同一レシートで複数回発行されるため、時分秒+マイクロ秒を含める。
    """
    dt = datetime.now()
    safe_label = str(label).replace("/", "_").replace("\\", "_")
    return f"{dt.strftime('%Y_%m_%d_%H_%M_%S')}_{dt.microsecond:06d}_{safe_label}_{receipt_no}.pdf"


def generate_checkout_pdf_for_items(
    receipt_no: str,
    paid_items: list[dict],
    subtotal: int,
    coupon_amount: int,
    discount_rate: int,
    discount_amount: int,
    total: int,
    manager_no: str,
    label: str = "会計",
    display_receipt_no: str | None = None,
) -> str:
    """
    個別会計・通常会計・合算内訳など、今回会計した商品だけをPDFにする。
    画面表示・保存ダイアログ表示は行わない。
    """
    PDF_DIR.mkdir(parents=True, exist_ok=True)

    with db() as conn:
        migrate_existing_db(conn)
        receipt = conn.execute(
            "SELECT * FROM receipts WHERE receipt_no = ?",
            (receipt_no,),
        ).fetchone()
        if receipt is None:
            raise ValueError("receipt not found")

    issue_no = next_daily_issue_no("checkout_receipt")
    pdf_datetime = format_pdf_datetime(now())
    pdf_file = PDF_DIR / checkout_pdf_filename_custom(receipt_no, label)

    shown_receipt_no = str(display_receipt_no or receipt_no)

    # 下部の見切れ防止のため、項目数だけでなく固定余白を十分に確保する。
    extra_lines = 2 + (1 if coupon_amount else 0) + (1 if discount_amount else 0)
    width = 80 * mm
    height = max(110 * mm, (60 + 7 * max(1, len(paid_items)) + 8 * extra_lines + 18) * mm)

    c = canvas.Canvas(str(pdf_file), pagesize=(width, height))
    w, h = width, height

    margin_x = 8 * mm
    y = h - 9 * mm

    c.setFont(PDF_FONT, 12.5)
    c.drawCentredString(w / 2, y, STORE_NAME)

    y -= 8 * mm
    c.setFont(PDF_FONT, 8.5)
    c.drawString(margin_x, y, f"No.{format_issue_no(issue_no)}")
    c.drawRightString(w - margin_x, y, pdf_datetime)

    y -= 7 * mm
    c.setFont(PDF_FONT, 9)
    c.drawString(margin_x, y, "責任者番号")
    c.drawRightString(w - margin_x, y, str(manager_no or "-"))

    y -= 9 * mm
    c.setFont(PDF_FONT, 9)
    c.drawString(margin_x, y, "レシート番号")

    y -= 7 * mm
    c.setFont(PDF_FONT, 13)
    c.drawString(margin_x + 7 * mm, y, shown_receipt_no)

    y -= 9 * mm
    c.line(margin_x, y, w - margin_x, y)
    y -= 7 * mm

    name_x = margin_x
    qty_x = w - 29 * mm
    amount_x = w - margin_x
    name_max_width = 36 * mm

    if not paid_items:
        c.setFont(PDF_FONT, 9)
        c.drawString(margin_x, y, "会計商品なし")
        y -= 7 * mm
    else:
        for item in paid_items:
            draw_limited_text(c, item["product_name"], name_x, y, name_max_width, PDF_FONT, 8.5)
            c.setFont(PDF_FONT, 8.5)
            c.drawRightString(qty_x, y, str(int(item["quantity"])))
            c.drawRightString(amount_x, y, f"{int(item['line_total']):,}円")
            y -= 7 * mm

    y -= 1 * mm
    c.line(margin_x, y, w - margin_x, y)
    y -= 7 * mm

    c.setFont(PDF_FONT, 8.5)
    c.drawString(margin_x, y, "小計")
    c.drawRightString(w - margin_x, y, f"{int(subtotal):,}円")

    if coupon_amount:
        y -= 6 * mm
        c.drawString(margin_x, y, "クーポン")
        c.drawRightString(w - margin_x, y, f"-{int(coupon_amount):,}円")

    if discount_amount:
        y -= 6 * mm
        c.drawString(margin_x, y, f"割引 {int(discount_rate)}%")
        c.drawRightString(w - margin_x, y, f"-{int(discount_amount):,}円")

    y -= 8 * mm
    c.setFont(PDF_FONT, 10)
    c.drawString(margin_x, y, "合計")
    c.setFont(PDF_FONT, 13)
    c.drawRightString(w - margin_x, y, f"{int(total):,}円")

    c.showPage()
    c.save()
    return str(pdf_file)


def generate_checkout_pdf(receipt_no: str) -> str:
    """
    会計確定時のレシートPDFをバックエンドで自動生成する。
    画面表示・保存ダイアログ表示は行わない。
    幅は固定、高さは商品数に応じて可変。
    """
    PDF_DIR.mkdir(parents=True, exist_ok=True)

    with db() as conn:
        migrate_existing_db(conn)
        receipt = conn.execute(
            "SELECT * FROM receipts WHERE receipt_no = ?",
            (receipt_no,),
        ).fetchone()
        if receipt is None:
            raise ValueError("receipt not found")

        items = conn.execute(
            """
            SELECT product_name, SUM(quantity) AS quantity, SUM(line_total) AS line_total
            FROM order_items
            WHERE receipt_no = ?
            GROUP BY product_name
            ORDER BY MIN(id)
            """,
            (receipt_no,),
        ).fetchall()

    paid_at = receipt["paid_at"] if "paid_at" in receipt.keys() else None
    pdf_datetime = format_pdf_datetime(paid_at)
    manager_no = receipt["manager_no"] if "manager_no" in receipt.keys() and receipt["manager_no"] else "-"
    issue_no = next_daily_issue_no("checkout_receipt")

    pdf_file = PDF_DIR / checkout_pdf_filename(receipt_no, paid_at)

    coupon_amount = int(receipt["coupon_amount"] or 0)
    discount_amount = int(receipt["discount_amount"] or 0)
    extra_lines = 2 + (1 if coupon_amount else 0) + (1 if discount_amount else 0)

    width = 80 * mm
    height = max(110 * mm, (60 + 7 * max(1, len(items)) + 8 * extra_lines + 18) * mm)

    c = canvas.Canvas(str(pdf_file), pagesize=(width, height))
    w, h = width, height

    margin_x = 8 * mm
    y = h - 9 * mm

    c.setFont(PDF_FONT, 12.5)
    c.drawCentredString(w / 2, y, STORE_NAME)

    y -= 8 * mm
    c.setFont(PDF_FONT, 8.5)
    c.drawString(margin_x, y, f"No.{format_issue_no(issue_no)}")
    c.drawRightString(w - margin_x, y, pdf_datetime)

    y -= 7 * mm
    c.setFont(PDF_FONT, 9)
    c.drawString(margin_x, y, "責任者番号")
    c.drawRightString(w - margin_x, y, str(manager_no))

    y -= 9 * mm
    c.setFont(PDF_FONT, 9)
    c.drawString(margin_x, y, "レシート番号")

    y -= 7 * mm
    c.setFont(PDF_FONT, 13)
    c.drawString(margin_x + 7 * mm, y, str(receipt["receipt_no"]))

    y -= 9 * mm
    c.line(margin_x, y, w - margin_x, y)
    y -= 7 * mm

    subtotal = 0
    name_x = margin_x
    qty_x = w - 29 * mm
    amount_x = w - margin_x
    name_max_width = 36 * mm

    if not items:
        c.setFont(PDF_FONT, 9)
        c.drawString(margin_x, y, "注文商品なし")
        y -= 7 * mm
    else:
        for item in items:
            name = item["product_name"]
            qty = int(item["quantity"] or 0)
            line_total = int(item["line_total"] or 0)
            subtotal += line_total

            draw_limited_text(c, name, name_x, y, name_max_width, PDF_FONT, 8.5)
            c.setFont(PDF_FONT, 8.5)
            c.drawRightString(qty_x, y, str(qty))
            c.drawRightString(amount_x, y, f"{line_total:,}円")
            y -= 7 * mm

    total = int(receipt["total_amount"] or max(0, subtotal - coupon_amount - discount_amount))

    y -= 1 * mm
    c.line(margin_x, y, w - margin_x, y)
    y -= 7 * mm

    c.setFont(PDF_FONT, 8.5)
    c.drawString(margin_x, y, "小計")
    c.drawRightString(w - margin_x, y, f"{subtotal:,}円")

    if coupon_amount:
        y -= 6 * mm
        c.drawString(margin_x, y, "クーポン")
        c.drawRightString(w - margin_x, y, f"-{coupon_amount:,}円")

    if discount_amount:
        y -= 6 * mm
        c.drawString(margin_x, y, f"割引 {int(receipt['discount_rate'] if 'discount_rate' in receipt.keys() and receipt['discount_rate'] else 0)}%")
        c.drawRightString(w - margin_x, y, f"-{discount_amount:,}円")

    y -= 8 * mm
    c.setFont(PDF_FONT, 10)
    c.drawString(margin_x, y, "合計")
    c.setFont(PDF_FONT, 13)
    c.drawRightString(w - margin_x, y, f"{total:,}円")

    c.showPage()
    c.save()

    rel = f"receipts/{pdf_file.name}"
    with db() as conn:
        migrate_existing_db(conn)
        conn.execute(
            "UPDATE receipts SET checkout_pdf_path = ? WHERE receipt_no = ?",
            (rel, receipt_no),
        )
    return str(pdf_file)

def get_table(table_no: int):
    with db() as conn:
        return conn.execute("SELECT * FROM tables WHERE table_no = ?", (table_no,)).fetchone()




def parse_payment_breakdown(raw: str) -> list[dict]:
    """
    レジ画面から送られる支払内訳JSONを検証して返す。
    例: [{"method":"電子マネー","amount":1000},{"method":"現金","amount":300}]
    """
    import json
    allowed = {"現金", "電子マネー", "クレジット", "QRコード", "WAON", "WAONポイント"}
    try:
        rows = json.loads(raw or "[]")
    except json.JSONDecodeError:
        return []

    if not isinstance(rows, list):
        return []

    result = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        method = str(row.get("method", "")).strip()
        amount = parse_yen_input(row.get("amount", 0))
        if method in allowed and amount > 0:
            result.append({"method": method, "amount": amount})
    return result


def summarize_payment_methods(payments: list[dict]) -> str:
    if not payments:
        return ""
    return " + ".join([f"{p['method']}:{int(p['amount'])}円" for p in payments])


def save_register_payments(conn: sqlite3.Connection, log_id: int | None, receipt_no: str, payments: list[dict]) -> None:
    for p in payments:
        conn.execute(
            """
            INSERT INTO register_payments(register_log_id, receipt_no, payment_method, amount, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (log_id, receipt_no, p["method"], int(p["amount"]), now()),
        )



def calc_unpaid_subtotal(receipt_no: str) -> int:
    """伝票内の未会計分だけを小計として計算する。"""
    with db() as conn:
        migrate_existing_db(conn)
        rows = conn.execute(
            """
            SELECT unit_price, quantity, COALESCE(paid_quantity, 0) AS paid_quantity
            FROM order_items
            WHERE receipt_no = ?
            """,
            (receipt_no,),
        ).fetchall()

        subtotal = 0
        for row in rows:
            remain_qty = max(0, int(row["quantity"] or 0) - int(row["paid_quantity"] or 0))
            subtotal += int(row["unit_price"] or 0) * remain_qty

        conn.execute(
            """
            UPDATE receipts
            SET subtotal = ?,
                total_amount = ?
            WHERE receipt_no = ?
            """,
            (subtotal, subtotal, receipt_no),
        )

    return subtotal


def receipt_has_unpaid_items(conn: sqlite3.Connection, receipt_no: str) -> bool:
    row = conn.execute(
        """
        SELECT COUNT(*) AS c
        FROM order_items
        WHERE receipt_no = ?
          AND quantity > COALESCE(paid_quantity, 0)
        """,
        (receipt_no,),
    ).fetchone()
    return int(row["c"] or 0) > 0


def mark_receipt_paid_if_complete(conn: sqlite3.Connection, receipt_no: str, manager_no: str | None = None, payment_method: str | None = None) -> bool:
    """全商品が会計済みなら伝票をpaidにする。未会計が残る場合はactiveのまま。"""
    has_unpaid = receipt_has_unpaid_items(conn, receipt_no)
    receipt = conn.execute("SELECT * FROM receipts WHERE receipt_no = ?", (receipt_no,)).fetchone()
    if receipt is None:
        return False

    if not has_unpaid:
        conn.execute(
            """
            UPDATE receipts
            SET status = 'paid',
                subtotal = 0,
                total_amount = 0,
                manager_no = COALESCE(?, manager_no),
                payment_method = COALESCE(?, payment_method),
                paid_at = COALESCE(paid_at, ?),
                voided_at = COALESCE(voided_at, ?)
            WHERE receipt_no = ?
            """,
            (manager_no, payment_method, now(), now(), receipt_no),
        )
        conn.execute("UPDATE tables SET status = 'paid' WHERE table_no = ?", (receipt["table_no"],))
        return True

    # 未会計分が残る場合は、残額だけをreceiptへ反映
    subtotal = 0
    rows = conn.execute(
        """
        SELECT unit_price, quantity, COALESCE(paid_quantity, 0) AS paid_quantity
        FROM order_items
        WHERE receipt_no = ?
        """,
        (receipt_no,),
    ).fetchall()
    for item in rows:
        remain_qty = max(0, int(item["quantity"] or 0) - int(item["paid_quantity"] or 0))
        subtotal += int(item["unit_price"] or 0) * remain_qty

    conn.execute(
        """
        UPDATE receipts
        SET status = 'active',
            subtotal = ?,
            total_amount = ?
        WHERE receipt_no = ?
        """,
        (subtotal, subtotal, receipt_no),
    )
    return False


def calc_subtotal(receipt_no: str) -> int:
    with db() as conn:
        total = conn.execute(
            "SELECT COALESCE(SUM(line_total), 0) FROM order_items WHERE receipt_no = ?",
            (receipt_no,),
        ).fetchone()[0]
        conn.execute("UPDATE receipts SET subtotal = ? WHERE receipt_no = ?", (total, receipt_no))
    return int(total)


def recalc_receipt_total(conn: sqlite3.Connection, receipt_no: str) -> int:
    """order_items.receipt_no を正として小計を再計算する互換関数。"""
    row = conn.execute(
        "SELECT COALESCE(SUM(line_total), 0) AS subtotal FROM order_items WHERE receipt_no = ?",
        (receipt_no,),
    ).fetchone()
    subtotal = int(row["subtotal"] if isinstance(row, sqlite3.Row) else row[0] or 0)
    conn.execute(
        """
        UPDATE receipts
        SET subtotal = ?,
            total_amount = MAX(0, ? - coupon_amount - discount_amount)
        WHERE receipt_no = ?
        """,
        (subtotal, subtotal, receipt_no),
    )
    return subtotal


def ensure_order_cancel_tables(conn: sqlite3.Connection) -> None:
    """注文キャンセル履歴を保存するテーブルを作成する。"""
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS order_cancel_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            receipt_no TEXT NOT NULL,
            table_no INTEGER NOT NULL,
            order_item_id INTEGER,
            product_name TEXT NOT NULL,
            unit_price INTEGER NOT NULL DEFAULT 0,
            canceled_quantity INTEGER NOT NULL DEFAULT 0,
            canceled_amount INTEGER NOT NULL DEFAULT 0,
            manager_no TEXT NOT NULL,
            reason TEXT,
            canceled_at TEXT NOT NULL
        )
        """
    )


def ensure_auth_tables(conn: sqlite3.Connection) -> None:
    """店員・管理者ログイン用のユーザーテーブルを作成する。"""
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS app_users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            display_name TEXT,
            role TEXT NOT NULL DEFAULT 'staff',
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT,
            retired_at TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS operation_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            actor_user_id INTEGER,
            actor_username TEXT,
            action TEXT NOT NULL,
            target_type TEXT,
            target_id TEXT,
            detail TEXT,
            created_at TEXT NOT NULL
        )
        """
    )


def normalize_username(username: str) -> str:
    return re.sub(r"\s+", "", str(username or "")).strip()


def valid_user_role(role: str) -> str:
    role = str(role or "staff").strip().lower()
    return role if role in {"admin", "staff"} else "staff"


def role_label(role: str) -> str:
    return {"admin": "管理者", "staff": "店員"}.get(str(role), "店員")


def user_store_path() -> Path:
    """一時運用用のID/PW/権限TXT。config.txt の USER_TXT_PATH があれば優先する。"""
    config = load_config()
    raw = str(config.get("USER_TXT_PATH", "")).strip()
    if raw:
        path = Path(raw)
        if not path.is_absolute():
            path = BASE_DIR / path
        return path
    return USER_TXT_PATH


def user_store_enabled() -> bool:
    """
    ユーザー情報は一時的に staff_users.txt を正として扱う。
    書式の先頭3項目は必ず ID:PW:権限。
    4項目目以降はアプリが状態管理用に使う。
    """
    return True


def parse_txt_user_line(line: str, line_no: int) -> dict | None:
    raw = line.strip()
    if not raw or raw.startswith("#"):
        return None
    parts = raw.split(":")
    if len(parts) < 3:
        return None
    username = normalize_username(parts[0])
    password = parts[1].strip()
    role = valid_user_role(parts[2])
    is_active = 1
    display_name = ""
    created_at = ""
    updated_at = ""
    if len(parts) >= 4 and parts[3].strip() in {"0", "1"}:
        is_active = int(parts[3].strip())
    if len(parts) >= 5:
        display_name = parts[4].strip()
    if len(parts) >= 6:
        created_at = parts[5].strip()
    if len(parts) >= 7:
        updated_at = parts[6].strip()
    return {
        "id": int(line_no),
        "line_no": int(line_no),
        "username": username,
        "password": password,
        "password_hash": "",
        "display_name": display_name,
        "role": role,
        "is_active": int(is_active),
        "created_at": created_at or "-",
        "updated_at": updated_at or "-",
        "retired_at": "" if int(is_active) else (updated_at or now()),
    }


def load_txt_users(include_inactive: bool = True) -> list[dict]:
    path = user_store_path()
    if not path.exists():
        return []
    users = []
    for line_no, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        user = parse_txt_user_line(line, line_no)
        if user is None:
            continue
        if include_inactive or int(user["is_active"]):
            users.append(user)
    return users


def serialize_txt_user(user: dict) -> str:
    """
    先頭3項目はユーザー要望どおり ID:PW:権限。
    4項目目以降は退会/復帰と一覧表示のための内部管理項目。
    """
    username = normalize_username(user.get("username", ""))
    password = str(user.get("password", "")).strip()
    role = valid_user_role(user.get("role", "staff"))
    is_active = "1" if int(user.get("is_active", 1) or 0) else "0"
    display_name = str(user.get("display_name", "")).replace(":", "").strip()
    created_at = str(user.get("created_at", "") or now()).replace(":", "-")
    updated_at = str(user.get("updated_at", "") or now()).replace(":", "-")
    return f"{username}:{password}:{role}:{is_active}:{display_name}:{created_at}:{updated_at}"


def save_txt_users(users: list[dict]) -> None:
    path = user_store_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    header = [
        "# 一時ユーザー管理ファイル",
        "# 書式: ID:PW:権限:状態:表示名:作成日:更新日",
        "# 権限: admin=管理者 / staff=店員",
        "# 状態: 1=有効 / 0=退会",
        "# 注意: PWは平文保存です。本番運用ではDB+ハッシュ方式に戻してください。",
    ]
    body = [serialize_txt_user(u) for u in users if normalize_username(u.get("username", ""))]
    path.write_text("\n".join(header + body) + "\n", encoding="utf-8")


def find_txt_user_by_username(username: str, active_only: bool = False) -> dict | None:
    target = normalize_username(username)
    for user in load_txt_users(include_inactive=not active_only):
        if normalize_username(user.get("username", "")) == target:
            if active_only and not int(user.get("is_active", 0)):
                return None
            return user
    return None


def find_txt_user_by_id(user_id: int, include_inactive: bool = True) -> dict | None:
    for user in load_txt_users(include_inactive=include_inactive):
        if int(user.get("id", -1)) == int(user_id):
            return user
    return None


def replace_txt_user(user_id: int, replacement: dict) -> bool:
    users = load_txt_users(include_inactive=True)
    replaced = False
    for i, user in enumerate(users):
        if int(user.get("id", -1)) == int(user_id):
            replacement["id"] = user["id"]
            replacement["line_no"] = user["line_no"]
            replacement.setdefault("created_at", user.get("created_at", ""))
            users[i] = replacement
            replaced = True
            break
    if replaced:
        save_txt_users(users)
    return replaced


def ensure_default_admin_user(conn: sqlite3.Connection | None = None) -> None:
    """初回起動時に staff_users.txt へ最低1人の管理者を作る。"""
    if conn is not None:
        ensure_auth_tables(conn)
    users = load_txt_users(include_inactive=True)
    if users:
        return

    config = load_config()
    username = normalize_username(
        config.get("INITIAL_ADMIN_ID")
        or config.get("ADMIN_ID")
        or config.get("ADMIN_USER")
        or "admin"
    )
    password = str(
        config.get("INITIAL_ADMIN_PASSWORD")
        or config.get("ADMIN_PASSWORD")
        or "admin1234"
    )
    user = {
        "id": 1,
        "line_no": 1,
        "username": username,
        "password": password,
        "display_name": "初期管理者",
        "role": "admin",
        "is_active": 1,
        "created_at": now().replace(":", "-"),
        "updated_at": now().replace(":", "-"),
    }
    save_txt_users([user])
    print("=" * 60)
    print("初期管理者ユーザーを staff_users.txt に作成しました。")
    print(f"ID: {username}")
    print("PW: config.txt の INITIAL_ADMIN_PASSWORD / ADMIN_PASSWORD、未設定時は admin1234")
    print("運用前に /staff → ユーザー管理 で必ず変更してください。")
    print("=" * 60)


def current_user() -> dict | None:
    username = session.get("username", "")
    user_id = session.get("user_id")
    if username:
        user = find_txt_user_by_username(str(username), active_only=True)
        if user is not None:
            return user
    if user_id:
        try:
            return find_txt_user_by_id(int(user_id), include_inactive=False)
        except (TypeError, ValueError):
            return None
    return None


def current_username() -> str:
    user = current_user()
    if user is None:
        return ""
    return str(user.get("username") or "")


def login_user(user: dict) -> None:
    session.clear()
    session["user_id"] = int(user["id"])
    session["username"] = str(user["username"])
    session["role"] = str(user["role"])
    session["logged_in_at"] = now()


def logout_user() -> None:
    session.clear()


def write_operation_log(
    conn: sqlite3.Connection,
    action: str,
    target_type: str = "",
    target_id: str = "",
    detail: str = "",
) -> None:
    ensure_auth_tables(conn)
    actor_id = session.get("user_id")
    actor_name = session.get("username", "")
    conn.execute(
        """
        INSERT INTO operation_logs(actor_user_id, actor_username, action, target_type, target_id, detail, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (actor_id, actor_name, action, target_type, target_id, detail, now()),
    )


def is_staff_authenticated() -> bool:
    return current_user() is not None


def is_staff_admin_authenticated() -> bool:
    user = current_user()
    return user is not None and str(user.get("role")) == "admin"


def require_staff_or_redirect(message: str = "店員ログインが必要です。"):
    """HTML画面・POST操作用の店員認証チェック。URL直打ちによる不正操作を防ぐ。"""
    if is_staff_authenticated():
        return None
    flash(message)
    return redirect(url_for("staff_portal"))


def require_staff_or_json(message: str = "店員ログインが必要です。"):
    """fetch/API用の店員認証チェック。未認証時はJSON 403を返す。"""
    if is_staff_authenticated():
        return None
    return jsonify({"ok": False, "message": message}), 403


def require_admin_or_redirect(message: str = "管理者権限が必要です。"):
    guard = require_staff_or_redirect("先に店員IDでログインしてください。")
    if guard:
        return guard
    if is_staff_admin_authenticated():
        return None
    flash(message)
    return redirect(url_for("staff_portal"))


def require_admin_or_json(message: str = "管理者権限が必要です。"):
    if not is_staff_authenticated():
        return jsonify({"ok": False, "message": "店員ログインが必要です。"}), 403
    if not is_staff_admin_authenticated():
        return jsonify({"ok": False, "message": message}), 403
    return None


def get_staff_pin() -> str:
    """旧互換用。現在はID/PWログインを使う。"""
    config = load_config()
    pin = config.get("STAFF_PIN", "") or config.get("STAFF_CANCEL_PIN", "") or "0000"
    return str(pin).strip() or "0000"


def get_staff_cancel_pin() -> str:
    return get_staff_pin()


def is_staff_cancel_authenticated() -> bool:
    return is_staff_authenticated()


def get_staff_admin_pin() -> str:
    config = load_config()
    pin = (
        config.get("STAFF_ADMIN_PIN", "")
        or config.get("ADMIN_PIN", "")
        or config.get("STAFF_PIN", "")
        or config.get("STAFF_CANCEL_PIN", "")
        or "0000"
    )
    return str(pin).strip() or "0000"


def active_cancellable_order_items() -> list[sqlite3.Row]:
    """会計前で、まだキャンセル可能な注文行を取得する。"""
    with db() as conn:
        migrate_existing_db(conn)
        ensure_order_cancel_tables(conn)
        return conn.execute(
            """
            SELECT
                oi.id AS item_id,
                r.receipt_no,
                r.table_no,
                oi.product_name,
                oi.unit_price,
                oi.quantity,
                COALESCE(oi.paid_quantity, 0) AS paid_quantity,
                (oi.quantity - COALESCE(oi.paid_quantity, 0)) AS cancellable_quantity,
                ((oi.quantity - COALESCE(oi.paid_quantity, 0)) * oi.unit_price) AS cancellable_amount,
                oi.created_at
            FROM order_items oi
            JOIN receipts r ON r.receipt_no = oi.receipt_no
            WHERE r.status = 'active'
              AND oi.quantity > COALESCE(oi.paid_quantity, 0)
            ORDER BY r.table_no, r.receipt_no, oi.created_at, oi.id
            """
        ).fetchall()


def recent_order_cancel_logs(limit: int = 30) -> list[sqlite3.Row]:
    with db() as conn:
        migrate_existing_db(conn)
        ensure_order_cancel_tables(conn)
        return conn.execute(
            """
            SELECT *
            FROM order_cancel_logs
            ORDER BY canceled_at DESC, id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()


def recalc_unpaid_subtotal_in_conn(conn: sqlite3.Connection, receipt_no: str) -> int:
    """キャンセル後に同一トランザクション内で未会計小計を再計算する。"""
    row = conn.execute(
        """
        SELECT COALESCE(SUM((quantity - COALESCE(paid_quantity, 0)) * unit_price), 0) AS subtotal
        FROM order_items
        WHERE receipt_no = ?
          AND quantity > COALESCE(paid_quantity, 0)
        """,
        (receipt_no,),
    ).fetchone()
    subtotal = int(row["subtotal"] or 0)
    conn.execute(
        """
        UPDATE receipts
        SET subtotal = ?,
            total_amount = MAX(0, ? - COALESCE(coupon_amount, 0) - COALESCE(discount_amount, 0))
        WHERE receipt_no = ?
        """,
        (subtotal, subtotal, receipt_no),
    )
    return subtotal


@app.route("/")
def root():
    return "OK: /staff または /entry または /register を直接開いてください。"



@app.route("/favicon.ico")
def favicon():
    return ("", 204)


@app.get("/photos/<path:filename>")
def serve_menu_photo(filename: str):
    """
    product_menu.txt で |001.png のように書いた商品写真を配信する。
    実ファイルはプロジェクト直下の photos/ に置く。
    例: BASE_DIR/photos/001.png -> /photos/001.png
    """
    return send_from_directory(PHOTO_DIR, filename)



def get_active_receipt_by_table(table_no: int):
    """
    指定テーブルの最新active伝票を返す。
    receiptsにid列が無いDBでも動くよう、created_at/receipt_noで並べる。
    """
    with db() as conn:
        migrate_existing_db(conn)
        return conn.execute(
            """
            SELECT *
            FROM receipts
            WHERE table_no = ? AND status = 'active'
            ORDER BY created_at DESC, receipt_no DESC
            LIMIT 1
            """,
            (table_no,),
        ).fetchone()


def get_receipt_by_no(receipt_no: str):
    """
    伝票番号から伝票を返す。
    同一receipt_noが存在する想定は薄いが、created_atで最新を取る。
    """
    with db() as conn:
        migrate_existing_db(conn)
        return conn.execute(
            """
            SELECT *
            FROM receipts
            WHERE receipt_no = ?
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (receipt_no,),
        ).fetchone()



@app.post("/entry/reissue")
def entry_reissue():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    """
    入店管理システム右上の再発行ポップアップ用。
    - QR再発行: テーブル番号または伝票番号からQR付き入店伝票を再発行
    - キッチン指示伝票: 伝票番号からキッチン伝票を再発行
    """
    reissue_type = request.form.get("reissue_type", "").strip()
    table_no_raw = request.form.get("table_no", "").strip()
    receipt_no = request.form.get("receipt_no", "").strip()

    if reissue_type == "qr":
        receipt = None

        if receipt_no:
            receipt = get_receipt_by_no(receipt_no)
        elif table_no_raw:
            try:
                table_no = int(table_no_raw)
            except ValueError:
                flash("テーブル番号は数値で入力してください。")
                return redirect(url_for("entry"))
            receipt = get_active_receipt_by_table(table_no)
        else:
            flash("QR再発行には、テーブル番号または伝票番号を入力してください。")
            return redirect(url_for("entry"))

        if receipt is None:
            flash("再発行対象の伝票が見つかりません。")
            return redirect(url_for("entry"))

        if receipt["status"] != "active":
            flash("会計済み・無効化済みのQRは再発行できません。")
            return redirect(url_for("entry"))

        # QR画像が削除済みでも再発行できるよう、既存tokenでQRを再生成する。
        qr_path = make_qr(receipt["receipt_no"], receipt["token"])
        with db() as conn:
            migrate_existing_db(conn)
            conn.execute("UPDATE receipts SET qr_path = ? WHERE receipt_no = ?", (qr_path, receipt["receipt_no"]))

        generate_ticket_pdf(receipt["receipt_no"])
        flash(f"QR付き入店伝票を再発行しました。伝票番号: {receipt['receipt_no']}")
        return redirect(url_for("entry"))

    if reissue_type == "kitchen":
        if not receipt_no:
            flash("キッチン指示伝票の再発行には、伝票番号を入力してください。")
            return redirect(url_for("entry"))

        table_no, submitted_items = build_submitted_items_from_receipt(receipt_no)
        if table_no is None or not submitted_items:
            flash("再発行できる注文データがありません。")
            return redirect(url_for("entry"))

        reissue_order_ticket_pdf_original(receipt_no)
        flash(f"キッチン指示伝票を原本と同じ表示で再発行しました。伝票番号: {receipt_no}")
        return redirect(url_for("entry"))

    flash("再発行種別を選択してください。")
    return redirect(url_for("entry"))


@app.route("/entry")
def entry():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    with db() as conn:
        tables = conn.execute("SELECT * FROM tables WHERE table_no BETWEEN 1 AND ? ORDER BY table_no", (TABLE_COUNT,)).fetchall()
    return render_template("entry.html", store_name=STORE_NAME, tables=tables)


@app.post("/entry/table/<int:table_no>")
def table_action(table_no: int):
    guard = require_staff_or_json()
    if guard:
        return guard
    row = get_table(table_no)
    if row is None:
        abort(404)

    status = row["status"]
    if status == "available":
        token = secrets.token_urlsafe(24)

        # migrate_existing_db() は列追加やUPDATEを行うため、同じ接続で直後に
        # BEGIN IMMEDIATE を開始すると SQLite の
        # "cannot start a transaction within a transaction" が発生する場合がある。
        # そのため、既存DB補正と入店処理トランザクションは接続を分ける。
        with db() as conn:
            migrate_existing_db(conn)

        with db() as conn:
            conn.execute("BEGIN IMMEDIATE")
            current = conn.execute("SELECT * FROM tables WHERE table_no = ?", (table_no,)).fetchone()
            if current is None:
                abort(404)
            if current["status"] != "available":
                return jsonify({
                    "ok": False,
                    "message": "このテーブルは別端末で処理済みです。画面を更新してください。",
                    "receipt_no": current["current_receipt_no"],
                }), 409

            receipt_no = next_receipt_no_in_conn(conn)
            qr_path = make_qr(receipt_no, token)
            conn.execute("""
                INSERT INTO receipts(receipt_no, table_no, token, qr_path, status, created_at)
                VALUES (?, ?, ?, ?, 'active', ?)
            """, (receipt_no, table_no, token, qr_path, now()))
            conn.execute("""
                UPDATE tables
                SET status = 'occupied', current_receipt_no = ?
                WHERE table_no = ? AND status = 'available'
            """, (receipt_no, table_no))
            write_operation_log(conn, "entry_create", "receipts", receipt_no, f"table_no={table_no}")
        generate_ticket_pdf(receipt_no)
        return jsonify({
            "ok": True,
            "action": "created",
            "receipt_no": receipt_no,
        })

    if status == "paid":
        with db() as conn:
            conn.execute("""
                UPDATE tables
                SET status = 'available', current_receipt_no = NULL
                WHERE table_no = ?
            """, (table_no,))
        return jsonify({"ok": True, "action": "reset"})

    if status == "occupied":
        return jsonify({
            "ok": False,
            "message": "このテーブルは使用中です。会計する場合はレジ画面を使用してください。",
            "receipt_no": row["current_receipt_no"],
        }), 409

    return jsonify({"ok": False, "message": "不明な状態です。"}), 400


@app.route("/receipt/<receipt_no>/pdf")
def receipt_pdf(receipt_no: str):
    guard = require_staff_or_redirect()
    if guard:
        return guard
    with db() as conn:
        receipt = conn.execute("SELECT * FROM receipts WHERE receipt_no = ?", (receipt_no,)).fetchone()
    if receipt is None:
        abort(404)
    if receipt["pdf_path"]:
        pdf_file = BASE_DIR / receipt["pdf_path"]
    else:
        pdf_file = Path(generate_ticket_pdf(receipt_no))
    if not pdf_file.exists():
        pdf_file = Path(generate_ticket_pdf(receipt_no))
    return send_file(pdf_file, mimetype="application/pdf", as_attachment=False, download_name=pdf_file.name)




@app.route("/receipt/<receipt_no>/preview")
def receipt_preview(receipt_no: str):
    guard = require_staff_or_redirect()
    if guard:
        return guard
    """PDFをブラウザに直接開かず、HTMLで伝票を表示する。
    Edge/ChromeのPDFダウンロード設定によるWindows保存ダイアログを避けるための画面。
    PDFファイル自体はreceiptsフォルダへ自動保存済み。
    """
    with db() as conn:
        receipt = conn.execute("SELECT * FROM receipts WHERE receipt_no = ?", (receipt_no,)).fetchone()
    if receipt is None:
        abort(404)
    if not receipt["pdf_path"]:
        generate_ticket_pdf(receipt_no)
        with db() as conn:
            receipt = conn.execute("SELECT * FROM receipts WHERE receipt_no = ?", (receipt_no,)).fetchone()
    pdf_name = Path(receipt["pdf_path"]).name if receipt["pdf_path"] else pdf_filename(receipt_no, receipt["created_at"])
    return render_template("ticket_preview.html", store_name=STORE_NAME, receipt=receipt, pdf_name=pdf_name)


def build_order_page_context(receipt: sqlite3.Row) -> dict:
    """モバイル注文画面・試験用メニュー画面で共通利用する表示用データを作る。"""
    products = add_i18n_to_products(load_product_menu())

    with db() as conn:
        migrate_existing_db(conn)
        fresh_receipt = conn.execute(
            "SELECT * FROM receipts WHERE receipt_no = ?",
            (receipt["receipt_no"],),
        ).fetchone()
        items = conn.execute(
            "SELECT * FROM order_items WHERE receipt_no = ? ORDER BY id DESC",
            (receipt["receipt_no"],),
        ).fetchall()
        subtotal = conn.execute(
            "SELECT COALESCE(SUM(line_total), 0) FROM order_items WHERE receipt_no = ?",
            (receipt["receipt_no"],),
        ).fetchone()[0]

    return {
        "store_name": STORE_NAME,
        "receipt": fresh_receipt or receipt,
        "products": products,
        "categories": build_category_view(products),
        "items": items,
        "subtotal": subtotal,
        "menu_display": load_menu_display_settings(),
        "order_name_i18n": order_name_i18n,
    }


def redirect_order_page(endpoint: str, endpoint_kwargs: dict | None = None):
    return redirect(url_for(endpoint, **(endpoint_kwargs or {})))


def submit_order_for_receipt(receipt: sqlite3.Row, redirect_endpoint: str, endpoint_kwargs: dict | None = None):
    """注文POST共通処理。本番QR注文と試験用メニュー表で処理を共有する。"""
    cart_raw = request.form.get("cart_json", "[]")
    try:
        cart = json.loads(cart_raw)
    except json.JSONDecodeError:
        cart = []

    if not isinstance(cart, list):
        cart = []

    submitted_items = []
    cart_errors = []
    lock_owner = f"order:{receipt['receipt_no']}:{request.remote_addr or 'unknown'}:{time.time()}"
    ok, locked_receipt = acquire_receipt_locks([receipt["receipt_no"]], lock_owner, ttl_seconds=45)
    if not ok:
        flash("ただいま店員が会計・取消処理中です。追加注文は店員にお声がけください。")
        return redirect_order_page(redirect_endpoint, endpoint_kwargs)

    try:
        with db() as conn:
            migrate_existing_db(conn)
            fresh_receipt = conn.execute(
                "SELECT * FROM receipts WHERE receipt_no = ?",
                (receipt["receipt_no"],),
            ).fetchone()
            if fresh_receipt is None or fresh_receipt["status"] != "active":
                flash("このQR/URLは無効、または会計済みです。")
                return redirect_order_page(redirect_endpoint, endpoint_kwargs)

            built_items, cart_errors = build_order_items_from_cart_with_errors(cart)
            submitted_items = insert_built_order_items(conn, receipt, built_items)

        if submitted_items:
            generate_order_ticket_pdf(
                receipt_no=receipt["receipt_no"],
                table_no=receipt["table_no"],
                submitted_items=submitted_items,
            )
            calc_subtotal(receipt["receipt_no"])
            flash("注文を送信しました。")
        else:
            if cart_errors:
                for error in cart_errors:
                    flash(error)
            else:
                flash("商品を1つ以上選択してください。セットメニューは追加項目の選択が必須です。")
    finally:
        release_receipt_locks([receipt["receipt_no"]], lock_owner)

    return redirect_order_page(redirect_endpoint, endpoint_kwargs)


@app.route("/mobile/<token>", methods=["GET", "POST"])
def mobile_order(token: str):
    with db() as conn:
        migrate_existing_db(conn)
        receipt = conn.execute("SELECT * FROM receipts WHERE token = ?", (token,)).fetchone()

    if receipt is None or receipt["status"] != "active":
        return render_template("mobile_disabled.html", store_name=STORE_NAME), 403

    # 初回アクセス時、まだ注文が一度も送信されていない場合は客数・子供用備品を入力させる。
    with db() as conn:
        migrate_existing_db(conn)
        order_count = conn.execute(
            "SELECT COUNT(*) AS c FROM order_items WHERE receipt_no = ?",
            (receipt["receipt_no"],),
        ).fetchone()["c"]

    if request.method == "POST" and request.form.get("form_type") == "guest_info":
        def to_non_negative_int(name: str) -> int:
            try:
                return max(0, int(request.form.get(name, "0") or 0))
            except ValueError:
                return 0

        adult_count = to_non_negative_int("adult_count")
        child_count = to_non_negative_int("child_count")
        child_chopsticks = to_non_negative_int("child_chopsticks")
        spoon_fork = to_non_negative_int("spoon_fork")

        if adult_count + child_count <= 0:
            flash("大人または子供の人数を1名以上入力してください。")
            return redirect(url_for("mobile_order", token=token))

        with db() as conn:
            migrate_existing_db(conn)
            conn.execute(
                """
                UPDATE receipts
                SET adult_count = ?,
                    child_count = ?,
                    child_chopsticks = ?,
                    spoon_fork = ?,
                    guest_info_submitted = 1
                WHERE receipt_no = ?
                """,
                (adult_count, child_count, child_chopsticks, spoon_fork, receipt["receipt_no"]),
            )

        generate_guest_info_pdf(
            receipt_no=receipt["receipt_no"],
            table_no=receipt["table_no"],
            adult_count=adult_count,
            child_count=child_count,
            child_chopsticks=child_chopsticks,
            spoon_fork=spoon_fork,
        )

        flash("人数情報を登録しました。続けて商品を選択してください。")
        return redirect(url_for("mobile_order", token=token))

    guest_info_done = int(receipt["guest_info_submitted"] if "guest_info_submitted" in receipt.keys() and receipt["guest_info_submitted"] is not None else 0)
    # 人数入力は guest_info_submitted で管理する。
    # 注文がまだ0件でも、一度登録済みなら再表示しない。
    if int(order_count or 0) == 0 and guest_info_done == 0:
        return render_template("mobile_guest_info.html", store_name=STORE_NAME, receipt=receipt)

    if request.method == "POST":
        return submit_order_for_receipt(receipt, "mobile_order", {"token": token})

    return render_template("mobile_order.html", **build_order_page_context(receipt))


@app.get("/api/receipt/<receipt_no>/amount")
def receipt_amount(receipt_no: str):
    guard = require_staff_or_json()
    if guard:
        return guard
    with db() as conn:
        migrate_existing_db(conn)
        receipt = conn.execute("SELECT * FROM receipts WHERE receipt_no = ?", (receipt_no,)).fetchone()
        if receipt is None or receipt["status"] != "active":
            return jsonify({"ok": False, "message": "有効なレシート番号が見つかりません。"}), 404
    subtotal = calc_unpaid_subtotal(receipt_no)
    return jsonify({"ok": True, "receipt_no": receipt_no, "table_no": receipt["table_no"], "subtotal": int(subtotal)})


def fetch_unpaid_item_rows(conn: sqlite3.Connection, receipt_no: str) -> list[sqlite3.Row]:
    """個別会計・合算会計のプレビュー用に未会計商品を取得する。"""
    return conn.execute(
        """
        SELECT
            oi.id,
            oi.receipt_no,
            oi.product_name,
            oi.unit_price,
            oi.quantity,
            COALESCE(oi.paid_quantity, 0) AS paid_quantity,
            (oi.quantity - COALESCE(oi.paid_quantity, 0)) AS remain_quantity,
            ((oi.quantity - COALESCE(oi.paid_quantity, 0)) * oi.unit_price) AS remain_total,
            oi.created_at
        FROM order_items oi
        WHERE oi.receipt_no = ?
          AND oi.quantity > COALESCE(oi.paid_quantity, 0)
        ORDER BY oi.created_at, oi.id
        """,
        (receipt_no,),
    ).fetchall()


@app.get("/api/receipt/<receipt_no>/unpaid-items")
def receipt_unpaid_items(receipt_no: str):
    """個別会計モーダルに表示する未会計商品一覧。"""
    guard = require_staff_or_json()
    if guard:
        return guard

    with db() as conn:
        migrate_existing_db(conn)
        receipt = conn.execute(
            "SELECT * FROM receipts WHERE receipt_no = ?",
            (receipt_no,),
        ).fetchone()
        if receipt is None or receipt["status"] != "active":
            return jsonify({"ok": False, "message": "有効なレシート番号が見つかりません。"}), 404

        rows = fetch_unpaid_item_rows(conn, receipt_no)

    return jsonify({
        "ok": True,
        "receipt_no": receipt_no,
        "items": [
            {
                "id": int(row["id"]),
                "product_name": row["product_name"],
                "unit_price": int(row["unit_price"] or 0),
                "quantity": int(row["quantity"] or 0),
                "paid_quantity": int(row["paid_quantity"] or 0),
                "remain_quantity": int(row["remain_quantity"] or 0),
                "line_total": int(row["remain_total"] or 0),
                "created_at": row["created_at"],
            }
            for row in rows
        ],
    })


@app.post("/api/register/partial-preview")
def register_partial_preview():
    """個別会計で選択した商品・数量だけの小計を返す。"""
    guard = require_staff_or_json()
    if guard:
        return guard

    data = request.get_json(silent=True) or {}
    receipt_no = str(data.get("receipt_no", "")).strip()
    selected_items = data.get("items", [])

    if not receipt_no:
        return jsonify({"ok": False, "message": "レシート番号がありません。"}), 400
    if not isinstance(selected_items, list) or not selected_items:
        return jsonify({"ok": False, "message": "会計する商品を1つ以上選択してください。"}), 400

    normalized = []
    subtotal = 0

    with db() as conn:
        migrate_existing_db(conn)
        receipt = conn.execute(
            "SELECT * FROM receipts WHERE receipt_no = ? AND status = 'active'",
            (receipt_no,),
        ).fetchone()
        if receipt is None:
            return jsonify({"ok": False, "message": "有効なレシート番号が見つかりません。"}), 404

        for selected in selected_items:
            if not isinstance(selected, dict):
                continue
            try:
                item_id = int(selected.get("id"))
                qty = int(selected.get("quantity", 0))
            except (TypeError, ValueError):
                continue
            if qty <= 0:
                continue

            item = conn.execute(
                """
                SELECT id, product_name, unit_price, quantity,
                       COALESCE(paid_quantity, 0) AS paid_quantity
                FROM order_items
                WHERE id = ? AND receipt_no = ?
                """,
                (item_id, receipt_no),
            ).fetchone()
            if item is None:
                continue

            remain_qty = max(0, int(item["quantity"] or 0) - int(item["paid_quantity"] or 0))
            pay_qty = min(qty, remain_qty)
            if pay_qty <= 0:
                continue

            unit_price = int(item["unit_price"] or 0)
            line_total = unit_price * pay_qty
            subtotal += line_total
            normalized.append({
                "id": int(item["id"]),
                "product_name": item["product_name"],
                "unit_price": unit_price,
                "quantity": pay_qty,
                "line_total": line_total,
                "remain_quantity": remain_qty,
            })

    if not normalized:
        return jsonify({"ok": False, "message": "個別会計できる未会計商品がありません。"}), 400

    return jsonify({
        "ok": True,
        "receipt_no": receipt_no,
        "items": normalized,
        "subtotal": int(subtotal),
    })


@app.post("/api/register/combined-preview")
def register_combined_preview():
    """合算会計で指定した複数伝票の未会計小計を返す。"""
    guard = require_staff_or_json()
    if guard:
        return guard

    data = request.get_json(silent=True) or {}
    raw_receipt_nos = data.get("receipt_nos", [])
    if not isinstance(raw_receipt_nos, list):
        return jsonify({"ok": False, "message": "レシート番号の指定が不正です。"}), 400

    receipt_nos = []
    for rn in raw_receipt_nos:
        s = str(rn or "").strip()
        if s and s not in receipt_nos:
            receipt_nos.append(s)

    if len(receipt_nos) < 2:
        return jsonify({"ok": False, "message": "合算会計は複数のレシート番号を指定してください。"}), 400

    result_receipts = []
    grand_total = 0

    with db() as conn:
        migrate_existing_db(conn)
        for rn in receipt_nos:
            receipt = conn.execute(
                "SELECT * FROM receipts WHERE receipt_no = ? AND status = 'active'",
                (rn,),
            ).fetchone()
            if receipt is None:
                continue

            rows = fetch_unpaid_item_rows(conn, rn)
            subtotal = sum(int(row["remain_total"] or 0) for row in rows)
            if subtotal <= 0:
                continue

            grand_total += subtotal
            result_receipts.append({
                "receipt_no": rn,
                "table_no": int(receipt["table_no"]),
                "subtotal": int(subtotal),
                "item_count": len(rows),
            })

    if not result_receipts:
        return jsonify({"ok": False, "message": "未会計データが見つかりません。"}), 404

    return jsonify({
        "ok": True,
        "receipts": result_receipts,
        "subtotal": int(grand_total),
    })




def receipt_numbers_for_lock(receipt_no: str, checkout_type: str, combined_receipts_raw: str) -> list[str]:
    if checkout_type == "合算会計":
        values = [x.strip() for x in re.split(r"[,\n\s]+", combined_receipts_raw or "") if x.strip()]
        if receipt_no and receipt_no not in values:
            values.insert(0, receipt_no)
        return list(dict.fromkeys(values))
    return [receipt_no] if receipt_no else []


def acquire_receipt_locks(receipt_nos: list[str], owner: str, ttl_seconds: int = 180) -> tuple[bool, str]:
    """複数端末・同時操作対策。会計・注文・取消処理中の伝票を短時間ロックする。

    migrate_existing_db() は既存DBへ列追加やUPDATEを行うため、同一接続内で
    BEGIN IMMEDIATE を後から実行すると SQLite の
    "cannot start a transaction within a transaction" が発生する場合がある。
    そのため、マイグレーションとロック取得は接続を分ける。
    """
    receipt_nos = [r for r in dict.fromkeys(receipt_nos) if r]
    if not receipt_nos:
        return True, ""

    # 既存DB補正はロック取得トランザクション外で完了させる。
    with db() as conn:
        migrate_existing_db(conn)

    now_ts = int(time.time())
    expires_at = now_ts + ttl_seconds
    with db() as conn:
        conn.execute("BEGIN IMMEDIATE")
        conn.execute("DELETE FROM receipt_locks WHERE expires_at < ?", (now_ts,))
        for rn in receipt_nos:
            existing = conn.execute("SELECT * FROM receipt_locks WHERE receipt_no = ?", (rn,)).fetchone()
            if existing and existing["locked_by"] != owner:
                return False, rn
        for rn in receipt_nos:
            conn.execute(
                """
                INSERT OR REPLACE INTO receipt_locks(receipt_no, locked_by, locked_at, expires_at)
                VALUES (?, ?, ?, ?)
                """,
                (rn, owner, now(), expires_at),
            )
    return True, ""


def release_receipt_locks(receipt_nos: list[str], owner: str) -> None:
    receipt_nos = [r for r in dict.fromkeys(receipt_nos) if r]
    if not receipt_nos:
        return
    with db() as conn:
        migrate_existing_db(conn)
        for rn in receipt_nos:
            conn.execute("DELETE FROM receipt_locks WHERE receipt_no = ? AND locked_by = ?", (rn, owner))




def ensure_daily_close_tables(conn: sqlite3.Connection) -> None:
    """営業開始時現金・日次締め結果を保存するテーブルを作成する。"""
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS daily_cash_openings (
            date_key TEXT PRIMARY KEY,
            opening_cash INTEGER NOT NULL DEFAULT 0,
            manager_no TEXT NOT NULL,
            note TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS daily_closings (
            date_key TEXT PRIMARY KEY,
            opening_cash INTEGER NOT NULL DEFAULT 0,
            expected_cash INTEGER NOT NULL DEFAULT 0,
            actual_cash INTEGER NOT NULL DEFAULT 0,
            difference_cash INTEGER NOT NULL DEFAULT 0,
            sales_total INTEGER NOT NULL DEFAULT 0,
            sales_count INTEGER NOT NULL DEFAULT 0,
            refund_total INTEGER NOT NULL DEFAULT 0,
            refund_count INTEGER NOT NULL DEFAULT 0,
            manager_no TEXT NOT NULL,
            note TEXT,
            pdf_path TEXT,
            closed_at TEXT NOT NULL
        )
        """
    )


def date_key_from_text(value: str | None = None) -> str:
    raw = str(value or "").strip()
    if raw:
        try:
            return datetime.strptime(raw, "%Y-%m-%d").strftime("%Y-%m-%d")
        except ValueError:
            pass
    return datetime.now().strftime("%Y-%m-%d")


def get_opening_cash(date_key: str) -> dict:
    with db() as conn:
        migrate_existing_db(conn)
        ensure_daily_close_tables(conn)
        row = conn.execute(
            "SELECT * FROM daily_cash_openings WHERE date_key = ?",
            (date_key,),
        ).fetchone()
    if row is None:
        return {"date_key": date_key, "opening_cash": 0, "manager_no": "", "note": "", "created_at": "", "updated_at": ""}
    return {
        "date_key": row["date_key"],
        "opening_cash": int(row["opening_cash"] or 0),
        "manager_no": row["manager_no"] or "",
        "note": row["note"] or "",
        "created_at": row["created_at"] or "",
        "updated_at": row["updated_at"] or "",
    }


def get_daily_closing(date_key: str) -> dict | None:
    with db() as conn:
        migrate_existing_db(conn)
        ensure_daily_close_tables(conn)
        row = conn.execute("SELECT * FROM daily_closings WHERE date_key = ?", (date_key,)).fetchone()
    if row is None:
        return None
    return {key: row[key] for key in row.keys()}


def upsert_opening_cash(date_key: str, opening_cash: int, manager_no: str, note: str = "") -> None:
    with db() as conn:
        migrate_existing_db(conn)
        ensure_daily_close_tables(conn)
        existing = conn.execute("SELECT date_key FROM daily_cash_openings WHERE date_key = ?", (date_key,)).fetchone()
        if existing is None:
            conn.execute(
                """
                INSERT INTO daily_cash_openings(date_key, opening_cash, manager_no, note, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (date_key, int(opening_cash), manager_no, note, now(), now()),
            )
        else:
            conn.execute(
                """
                UPDATE daily_cash_openings
                SET opening_cash = ?, manager_no = ?, note = ?, updated_at = ?
                WHERE date_key = ?
                """,
                (int(opening_cash), manager_no, note, now(), date_key),
            )
        write_operation_log(conn, "opening_cash_upsert", "daily_cash_openings", date_key, f"opening_cash={opening_cash}, manager_no={manager_no}")


def cash_drawer_rows_for_day(conn: sqlite3.Connection, date_key: str) -> list[sqlite3.Row]:
    start = f"{date_key} 00:00:00"
    end = f"{date_key} 23:59:59"
    return conn.execute(
        """
        SELECT action_type, amount, manager_no, note, created_at
        FROM cash_drawer_logs
        WHERE created_at BETWEEN ? AND ?
        ORDER BY created_at, id
        """,
        (start, end),
    ).fetchall()


def register_log_rows_for_day(conn: sqlite3.Connection, date_key: str) -> list[sqlite3.Row]:
    start = f"{date_key} 00:00:00"
    end = f"{date_key} 23:59:59"
    return conn.execute(
        """
        SELECT receipt_no, manager_no, checkout_type, payment_method, subtotal,
               coupon_amount, discount_rate, discount_amount, total_amount, created_at
        FROM register_logs
        WHERE created_at BETWEEN ? AND ?
        ORDER BY created_at, id
        """,
        (start, end),
    ).fetchall()


def order_cancel_rows_for_day(conn: sqlite3.Connection, date_key: str) -> list[sqlite3.Row]:
    start = f"{date_key} 00:00:00"
    end = f"{date_key} 23:59:59"
    if not table_exists(conn, "order_cancel_logs"):
        return []
    return conn.execute(
        """
        SELECT receipt_no, table_no, product_name, unit_price, canceled_quantity,
               canceled_amount, manager_no, reason, canceled_at
        FROM order_cancel_logs
        WHERE canceled_at BETWEEN ? AND ?
        ORDER BY canceled_at, id
        """,
        (start, end),
    ).fetchall()


def draw_report_text(c: canvas.Canvas, text: str, x: float, y: float, max_width: float, font_size: float = 8.5) -> None:
    draw_limited_text(c, str(text), x, y, max_width, PDF_FONT, font_size)


def daily_report_filename(date_key: str) -> str:
    safe_date = str(date_key).replace("-", "_")
    return f"{safe_date}_日次締め.pdf"


def generate_daily_close_pdf(date_key: str, summary: dict, actual_cash: int, manager_no: str, note: str = "") -> str:
    """日次締めPDFを生成し、daily_reports フォルダへ保存する。"""
    DAILY_REPORT_DIR.mkdir(parents=True, exist_ok=True)
    pdf_file = DAILY_REPORT_DIR / daily_report_filename(date_key)

    expected_cash = int(summary.get("cash_balance", 0) or 0)
    opening_cash = int(summary.get("opening_cash", 0) or 0)
    difference = int(actual_cash) - expected_cash

    with db() as conn:
        migrate_existing_db(conn)
        ensure_daily_close_tables(conn)
        register_logs = register_log_rows_for_day(conn, date_key)
        cash_rows = cash_drawer_rows_for_day(conn, date_key)
        cancel_rows = order_cancel_rows_for_day(conn, date_key)

    width = 210 * mm
    height = 297 * mm
    c = canvas.Canvas(str(pdf_file), pagesize=(width, height))
    margin = 12 * mm
    y = height - margin

    def new_page_if_needed(required_mm: float = 16) -> None:
        nonlocal y
        if y < margin + required_mm * mm:
            c.showPage()
            y = height - margin

    def title(text: str) -> None:
        nonlocal y
        new_page_if_needed(18)
        c.setFont(PDF_FONT, 14)
        c.drawString(margin, y, text)
        y -= 8 * mm

    def section(text: str) -> None:
        nonlocal y
        new_page_if_needed(14)
        y -= 2 * mm
        c.setFont(PDF_FONT, 11)
        c.drawString(margin, y, text)
        y -= 5 * mm
        c.line(margin, y, width - margin, y)
        y -= 6 * mm

    def line(left: str, right: str = "") -> None:
        nonlocal y
        new_page_if_needed(7)
        c.setFont(PDF_FONT, 9)
        c.drawString(margin, y, left)
        if right:
            c.drawRightString(width - margin, y, right)
        y -= 6 * mm

    def table_header(headers: list[str], xs: list[float]) -> None:
        nonlocal y
        new_page_if_needed(10)
        c.setFont(PDF_FONT, 8.5)
        for h, x in zip(headers, xs):
            c.drawString(x, y, h)
        y -= 4 * mm
        c.line(margin, y, width - margin, y)
        y -= 5 * mm

    title(f"{STORE_NAME} 日次締め報告書")
    line("対象日", date_key)
    line("締め日時", now())
    line("担当者", manager_no or "-")
    if note:
        line("メモ", note[:80])

    section("1. 売上サマリー")
    line("売上件数", f"{int(summary.get('count', 0)):,}件")
    line("売上金額", f"{int(summary.get('total', 0)):,}円")
    line("返金件数", f"{int(summary.get('refund_count', 0)):,}件")
    line("返金金額", f"{int(summary.get('refund_total', 0)):,}円")
    line("クーポン", f"{int(summary.get('coupon_count', 0)):,}件 / {int(summary.get('coupon_amount', 0)):,}円")

    section("2. 現金残高")
    line("営業開始時現金", f"{opening_cash:,}円")
    line("現金売上", f"{int(summary.get('cash_payment', 0)):,}円")
    line("レジ金入金", f"{int(summary.get('cash_deposit', 0)):,}円")
    line("レジ金回収・出金", f"{int(summary.get('cash_withdrawal', 0)):,}円")
    line("システム上の残高", f"{expected_cash:,}円")
    line("実際の金額", f"{int(actual_cash):,}円")
    line("差額", f"{difference:+,}円")

    section("3. 支払方法別")
    table_header(["支払方法", "回数", "金額"], [margin, margin + 55 * mm, margin + 90 * mm])
    if summary.get("payments"):
        for p in summary.get("payments", []):
            new_page_if_needed(7)
            c.setFont(PDF_FONT, 8.5)
            draw_report_text(c, p.get("payment_method", ""), margin, y, 50 * mm)
            c.drawRightString(margin + 70 * mm, y, f"{int(p.get('count', 0)):,}")
            c.drawRightString(margin + 110 * mm, y, f"{int(p.get('amount', 0)):,}円")
            y -= 6 * mm
    else:
        line("該当期間の支払データはありません。")

    section("4. 割引")
    line("クーポン", f"{int(summary.get('coupon_count', 0)):,}件 / {int(summary.get('coupon_amount', 0)):,}円")
    for d in summary.get("discounts", []):
        line(f"割引 {int(d.get('rate', 0))}%", f"{int(d.get('count', 0)):,}件 / {int(d.get('amount', 0)):,}円")

    section("5. レジ金処理明細")
    table_header(["時刻", "区分", "金額", "担当", "メモ"], [margin, margin + 34 * mm, margin + 58 * mm, margin + 88 * mm, margin + 112 * mm])
    if cash_rows:
        for r in cash_rows:
            new_page_if_needed(7)
            c.setFont(PDF_FONT, 8)
            draw_report_text(c, str(r["created_at"])[11:19], margin, y, 30 * mm, 8)
            draw_report_text(c, r["action_type"], margin + 34 * mm, y, 20 * mm, 8)
            c.drawRightString(margin + 80 * mm, y, f"{int(r['amount'] or 0):,}円")
            draw_report_text(c, r["manager_no"], margin + 88 * mm, y, 22 * mm, 8)
            draw_report_text(c, r["note"] or "", margin + 112 * mm, y, 70 * mm, 8)
            y -= 6 * mm
    else:
        line("レジ金処理はありません。")

    section("6. 会計ログ")
    table_header(["時刻", "伝票", "区分", "支払方法", "合計"], [margin, margin + 28 * mm, margin + 52 * mm, margin + 84 * mm, margin + 145 * mm])
    if register_logs:
        for r in register_logs:
            new_page_if_needed(7)
            c.setFont(PDF_FONT, 7.5)
            draw_report_text(c, str(r["created_at"])[11:19], margin, y, 24 * mm, 7.5)
            draw_report_text(c, r["receipt_no"], margin + 28 * mm, y, 22 * mm, 7.5)
            draw_report_text(c, r["checkout_type"] or "", margin + 52 * mm, y, 30 * mm, 7.5)
            draw_report_text(c, r["payment_method"] or "", margin + 84 * mm, y, 58 * mm, 7.5)
            c.drawRightString(width - margin, y, f"{int(r['total_amount'] or 0):,}円")
            y -= 5.4 * mm
    else:
        line("会計ログはありません。")

    section("7. 注文キャンセル")
    if cancel_rows:
        total_cancel_amount = sum(int(r["canceled_amount"] or 0) for r in cancel_rows)
        total_cancel_qty = sum(int(r["canceled_quantity"] or 0) for r in cancel_rows)
        line("取消件数", f"{len(cancel_rows):,}件")
        line("取消数量", f"{total_cancel_qty:,}個")
        line("取消金額", f"{total_cancel_amount:,}円")
    else:
        line("注文キャンセルはありません。")

    c.showPage()
    c.save()
    return str(pdf_file)


def save_daily_closing(date_key: str, summary: dict, actual_cash: int, manager_no: str, note: str, pdf_path: str) -> None:
    expected_cash = int(summary.get("cash_balance", 0) or 0)
    opening_cash = int(summary.get("opening_cash", 0) or 0)
    difference = int(actual_cash) - expected_cash
    rel_path = ""
    try:
        rel_path = str(Path(pdf_path).relative_to(BASE_DIR)).replace("\\", "/")
    except Exception:
        rel_path = str(pdf_path)

    with db() as conn:
        migrate_existing_db(conn)
        ensure_daily_close_tables(conn)
        conn.execute(
            """
            INSERT INTO daily_closings(
                date_key, opening_cash, expected_cash, actual_cash, difference_cash,
                sales_total, sales_count, refund_total, refund_count,
                manager_no, note, pdf_path, closed_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(date_key) DO UPDATE SET
                opening_cash=excluded.opening_cash,
                expected_cash=excluded.expected_cash,
                actual_cash=excluded.actual_cash,
                difference_cash=excluded.difference_cash,
                sales_total=excluded.sales_total,
                sales_count=excluded.sales_count,
                refund_total=excluded.refund_total,
                refund_count=excluded.refund_count,
                manager_no=excluded.manager_no,
                note=excluded.note,
                pdf_path=excluded.pdf_path,
                closed_at=excluded.closed_at
            """,
            (
                date_key, opening_cash, expected_cash, int(actual_cash), difference,
                int(summary.get("total", 0) or 0), int(summary.get("count", 0) or 0),
                int(summary.get("refund_total", 0) or 0), int(summary.get("refund_count", 0) or 0),
                manager_no, note, rel_path, now(),
            ),
        )
        write_operation_log(conn, "daily_close", "daily_closings", date_key, f"expected={expected_cash}, actual={actual_cash}, diff={difference}")

def get_date_range_from_request(default_today: bool = True) -> tuple[str, str]:
    today = datetime.now().strftime("%Y-%m-%d")
    date_from = request.args.get("date_from", "").strip() or (today if default_today else "1900-01-01")
    date_to = request.args.get("date_to", "").strip() or today
    return date_from, date_to



def calc_cash_balance(conn: sqlite3.Connection, date_from: str, date_to: str) -> int:
    """
    レジ内現金残高を計算する。
    基本式:
      営業開始時現金 + レジ金入金 + 現金支払 - レジ金回収
    ※返金は register_payments の元支払方法別差し引きロジックで現金支払額から差し引かれる。
    """
    start = f"{date_from} 00:00:00"
    end = f"{date_to} 23:59:59"

    opening_cash = 0
    # 単日の場合だけ営業開始時現金を加算する。範囲集計では各日の開始残高を足し上げる。
    rows = conn.execute(
        """
        SELECT COALESCE(SUM(opening_cash), 0) AS opening_cash
        FROM daily_cash_openings
        WHERE date_key BETWEEN ? AND ?
        """,
        (date_from, date_to),
    ).fetchone() if table_exists(conn, "daily_cash_openings") else None
    if rows is not None:
        opening_cash = int(rows["opening_cash"] or 0)

    cash_payment_row = conn.execute(
        """
        SELECT COALESCE(SUM(amount), 0) AS amount
        FROM register_payments
        WHERE created_at BETWEEN ? AND ?
          AND payment_method = '現金'
        """,
        (start, end),
    ).fetchone()
    cash_payment = int(cash_payment_row["amount"] or 0)

    deposit_row = conn.execute(
        """
        SELECT COALESCE(SUM(amount), 0) AS amount
        FROM cash_drawer_logs
        WHERE created_at BETWEEN ? AND ?
          AND action_type = '入金'
        """,
        (start, end),
    ).fetchone()
    deposit = int(deposit_row["amount"] or 0)

    withdrawal_row = conn.execute(
        """
        SELECT COALESCE(SUM(amount), 0) AS amount
        FROM cash_drawer_logs
        WHERE created_at BETWEEN ? AND ?
          AND action_type IN ('回収', '出金')
        """,
        (start, end),
    ).fetchone()
    withdrawal = int(withdrawal_row["amount"] or 0)

    return opening_cash + deposit + cash_payment - withdrawal



def get_sales_summary(date_from: str, date_to: str) -> dict:
    """
    売上・支払方法・クーポン/割引の集計。
    返金は「返金」行として独立表示せず、返金対象伝票の元の支払方法から自動的に差し引く。
    """
    start = f"{date_from} 00:00:00"
    end = f"{date_to} 23:59:59"

    with db() as conn:
        migrate_existing_db(conn)
        ensure_daily_close_tables(conn)

        total_row = conn.execute(
            """
            SELECT COALESCE(SUM(total_amount), 0) AS total, COUNT(*) AS count
            FROM register_logs
            WHERE created_at BETWEEN ? AND ?
              AND checkout_type NOT LIKE '%返金%'
            """,
            (start, end),
        ).fetchone()

        refund_row = conn.execute(
            """
            SELECT COALESCE(SUM(ABS(total_amount)), 0) AS total, COUNT(*) AS count
            FROM register_logs
            WHERE created_at BETWEEN ? AND ?
              AND checkout_type LIKE '%返金%'
            """,
            (start, end),
        ).fetchone()

        # 通常の支払方法別集計。返金行はここでは除外する。
        payment_rows = conn.execute(
            """
            SELECT payment_method,
                   COUNT(*) AS count,
                   COALESCE(SUM(amount), 0) AS amount
            FROM register_payments
            WHERE created_at BETWEEN ? AND ?
              AND payment_method <> '返金'
              AND amount > 0
            GROUP BY payment_method
            """,
            (start, end),
        ).fetchall()

        payments_map: dict[str, dict] = {}
        for row in payment_rows:
            method = row["payment_method"] or "不明"
            payments_map[method] = {
                "payment_method": method,
                "count": int(row["count"] or 0),
                "amount": int(row["amount"] or 0),
            }

        # 返金分は、返金対象伝票の直近会計ログに紐づく支払内訳から差し引く。
        refund_logs = conn.execute(
            """
            SELECT id, receipt_no, ABS(total_amount) AS refund_amount
            FROM register_logs
            WHERE created_at BETWEEN ? AND ?
              AND checkout_type LIKE '%返金%'
            ORDER BY created_at
            """,
            (start, end),
        ).fetchall()

        for refund in refund_logs:
            receipt_no = refund["receipt_no"]
            refund_amount = int(refund["refund_amount"] or 0)

            paid_log = conn.execute(
                """
                SELECT id
                FROM register_logs
                WHERE receipt_no = ?
                  AND checkout_type IN ('通常会計', '個別会計', '合算会計')
                  AND created_at <= (
                      SELECT created_at FROM register_logs WHERE id = ?
                  )
                ORDER BY created_at DESC
                LIMIT 1
                """,
                (receipt_no, refund["id"]),
            ).fetchone()

            if paid_log is None:
                continue

            original_payments = conn.execute(
                """
                SELECT payment_method, amount
                FROM register_payments
                WHERE register_log_id = ?
                  AND amount > 0
                ORDER BY id
                """,
                (paid_log["id"],),
            ).fetchall()

            remaining = refund_amount
            for p in original_payments:
                if remaining <= 0:
                    break
                method = p["payment_method"] or "不明"
                original_amount = int(p["amount"] or 0)
                deduction = min(original_amount, remaining)
                remaining -= deduction

                if method not in payments_map:
                    payments_map[method] = {
                        "payment_method": method,
                        "count": 0,
                        "amount": 0,
                    }
                payments_map[method]["amount"] -= deduction

                # 全額返金された支払方法は、回数も1回分戻す。
                if deduction >= original_amount:
                    payments_map[method]["count"] -= 1

        payments = [
            p for p in payments_map.values()
            if int(p["count"] or 0) != 0 or int(p["amount"] or 0) != 0
        ]
        payments.sort(key=lambda x: int(x["amount"] or 0), reverse=True)

        coupon_row = conn.execute(
            """
            SELECT COUNT(*) AS count,
                   COALESCE(SUM(coupon_amount), 0) AS amount
            FROM register_logs
            WHERE created_at BETWEEN ? AND ?
              AND checkout_type NOT LIKE '%返金%'
              AND COALESCE(coupon_amount, 0) > 0
            """,
            (start, end),
        ).fetchone()

        discount_rows = conn.execute(
            """
            SELECT discount_rate,
                   COUNT(*) AS count,
                   COALESCE(SUM(discount_amount), 0) AS amount
            FROM register_logs
            WHERE created_at BETWEEN ? AND ?
              AND checkout_type NOT LIKE '%返金%'
              AND COALESCE(discount_rate, 0) IN (5, 10, 20)
              AND COALESCE(discount_amount, 0) > 0
            GROUP BY discount_rate
            """,
            (start, end),
        ).fetchall()

        discounts = {
            5: {"rate": 5, "count": 0, "amount": 0},
            10: {"rate": 10, "count": 0, "amount": 0},
            20: {"rate": 20, "count": 0, "amount": 0},
        }
        for row in discount_rows:
            rate = int(row["discount_rate"] or 0)
            if rate in discounts:
                discounts[rate] = {
                    "rate": rate,
                    "count": int(row["count"] or 0),
                    "amount": int(row["amount"] or 0),
                }

        cash = conn.execute(
            """
            SELECT action_type, COALESCE(SUM(amount), 0) AS amount, COUNT(*) AS count
            FROM cash_drawer_logs
            WHERE created_at BETWEEN ? AND ?
            GROUP BY action_type
            """,
            (start, end),
        ).fetchall()

        opening_cash_row = conn.execute(
            """
            SELECT COALESCE(SUM(opening_cash), 0) AS amount
            FROM daily_cash_openings
            WHERE date_key BETWEEN ? AND ?
            """,
            (date_from, date_to),
        ).fetchone()
        opening_cash = int(opening_cash_row["amount"] or 0)

        cash_payment_row = conn.execute(
            """
            SELECT COALESCE(SUM(amount), 0) AS amount
            FROM register_payments
            WHERE created_at BETWEEN ? AND ?
              AND payment_method = '現金'
            """,
            (start, end),
        ).fetchone()
        cash_payment = int(cash_payment_row["amount"] or 0)

        deposit_row = conn.execute(
            """
            SELECT COALESCE(SUM(amount), 0) AS amount
            FROM cash_drawer_logs
            WHERE created_at BETWEEN ? AND ?
              AND action_type = '入金'
            """,
            (start, end),
        ).fetchone()
        cash_deposit = int(deposit_row["amount"] or 0)

        withdrawal_row = conn.execute(
            """
            SELECT COALESCE(SUM(amount), 0) AS amount
            FROM cash_drawer_logs
            WHERE created_at BETWEEN ? AND ?
              AND action_type IN ('回収', '出金')
            """,
            (start, end),
        ).fetchone()
        cash_withdrawal = int(withdrawal_row["amount"] or 0)

    return {
        "total": int(total_row["total"] or 0),
        "count": int(total_row["count"] or 0),
        "refund_total": int(refund_row["total"] or 0),
        "refund_count": int(refund_row["count"] or 0),
        "payments": payments,
        "coupon_count": int(coupon_row["count"] or 0),
        "coupon_amount": int(coupon_row["amount"] or 0),
        "discounts": [discounts[5], discounts[10], discounts[20]],
        "cash": cash,
        "opening_cash": opening_cash,
        "cash_payment": cash_payment,
        "cash_deposit": cash_deposit,
        "cash_withdrawal": cash_withdrawal,
        "cash_balance": calc_cash_balance(conn, date_from, date_to),
    }



def menu_category_map() -> dict[str, str]:
    mapping = {}
    try:
        for p in load_product_menu(include_inactive=True):
            mapping[p["name"]] = p["category"]
    except Exception:
        pass
    return mapping


def infer_category(product_name: str, mapping: dict[str, str]) -> str:
    base = str(product_name).split(" / ", 1)[0]
    return mapping.get(base, "未分類")


@app.route("/staff")
def staff_portal():
    """店員用ポータル。現場で使う各画面への入口を集約する。"""
    return render_template(
        "staff_portal.html",
        store_name=STORE_NAME,
        staff_authenticated=is_staff_authenticated(),
        admin_authenticated=is_staff_admin_authenticated(),
        user=current_user(),
        role_label=role_label,
    )


@app.route("/staff/test-menu", methods=["GET", "POST"])
def staff_test_menu():
    """スタッフ用：QRを読まずに、テーブル999固定の試験用メニュー表を開く。"""
    guard = require_staff_or_redirect()
    if guard:
        return guard

    receipt = get_or_create_test_menu_receipt()

    if request.method == "POST":
        return submit_order_for_receipt(receipt, "staff_test_menu")

    return render_template("mobile_order_test.html", **build_order_page_context(receipt))


@app.post("/staff/login")
def staff_login():
    username = normalize_username(request.form.get("username", ""))
    password = request.form.get("password", "")
    if not username or not password:
        flash("IDとパスワードを入力してください。")
        return redirect(url_for("staff_portal"))

    ensure_default_admin_user(None)
    user = find_txt_user_by_username(username, active_only=True)

    if user is not None and str(user.get("password", "")) == str(password):
        login_user(user)
        flash("スタッフメニューにログインしました。")
    else:
        flash("IDまたはパスワードが正しくありません。")
    return redirect(url_for("staff_portal"))


@app.post("/staff/logout")
def staff_logout():
    logout_user()
    flash("ログアウトしました。")
    return redirect(url_for("staff_portal"))


@app.post("/staff/admin-login")
def staff_admin_login():
    guard = require_staff_or_redirect("先にID/PWでログインしてください。")
    if guard:
        return guard
    if is_staff_admin_authenticated():
        flash("管理者としてログイン中です。")
    else:
        flash("管理者機能を使うには、管理者権限のユーザーでログインしてください。")
    return redirect(url_for("staff_portal"))


@app.post("/staff/admin-logout")
def staff_admin_logout():
    flash("管理者機能を終了するにはログアウトしてください。")
    return redirect(url_for("staff_portal"))


@app.route("/admin/users")
def admin_users_page():
    guard = require_admin_or_redirect("ユーザー管理は管理者権限が必要です。")
    if guard:
        return guard
    ensure_default_admin_user(None)
    with db() as conn:
        ensure_auth_tables(conn)
        logs = conn.execute(
            """
            SELECT * FROM operation_logs
            WHERE action LIKE 'user_%'
            ORDER BY created_at DESC, id DESC
            LIMIT 30
            """
        ).fetchall()
    users = sorted(load_txt_users(include_inactive=True), key=lambda u: (not int(u.get("is_active", 0)), str(u.get("role", "")), str(u.get("username", ""))))
    return render_template(
        "admin_users.html",
        store_name=STORE_NAME,
        users=users,
        logs=logs,
        current_user=current_user(),
        role_label=role_label,
    )


@app.post("/admin/users/create")
def admin_users_create():
    guard = require_admin_or_redirect("ユーザー登録は管理者権限が必要です。")
    if guard:
        return guard

    username = normalize_username(request.form.get("username", ""))
    password = request.form.get("password", "")
    display_name = request.form.get("display_name", "").strip().replace(":", "")
    role = valid_user_role(request.form.get("role", "staff"))

    if not username or not password:
        flash("登録にはIDと初期パスワードが必要です。")
        return redirect(url_for("admin_users_page"))
    if ":" in username or ":" in password:
        flash("IDとパスワードにコロン(:)は使えません。")
        return redirect(url_for("admin_users_page"))
    if len(password) < 8:
        flash("パスワードは8文字以上にしてください。")
        return redirect(url_for("admin_users_page"))
    if find_txt_user_by_username(username, active_only=False) is not None:
        flash("同じIDのユーザーが既に存在します。")
        return redirect(url_for("admin_users_page"))

    users = load_txt_users(include_inactive=True)
    user = {
        "id": len(users) + 1,
        "line_no": len(users) + 1,
        "username": username,
        "password": password,
        "display_name": display_name,
        "role": role,
        "is_active": 1,
        "created_at": now().replace(":", "-"),
        "updated_at": now().replace(":", "-"),
    }
    users.append(user)
    save_txt_users(users)
    with db() as conn:
        write_operation_log(conn, "user_create", "staff_users.txt", username, f"role={role}")
    flash("ユーザーを staff_users.txt に登録しました。")
    return redirect(url_for("admin_users_page"))


@app.post("/admin/users/<int:user_id>/update")
def admin_users_update(user_id: int):
    guard = require_admin_or_redirect("ユーザー変更は管理者権限が必要です。")
    if guard:
        return guard

    me = current_user()
    if me is None:
        flash("ログインし直してください。")
        return redirect(url_for("staff_portal"))

    username = normalize_username(request.form.get("username", ""))
    display_name = request.form.get("display_name", "").strip().replace(":", "")
    new_password = request.form.get("password", "")
    new_role = valid_user_role(request.form.get("role", "staff"))
    try:
        new_is_active = 1 if int(request.form.get("is_active", "1") or 1) else 0
    except ValueError:
        new_is_active = 1

    if not username:
        flash("IDは空にできません。")
        return redirect(url_for("admin_users_page"))
    if ":" in username or ":" in new_password:
        flash("IDとパスワードにコロン(:)は使えません。")
        return redirect(url_for("admin_users_page"))
    if new_password and len(new_password) < 8:
        flash("新しいパスワードは8文字以上にしてください。")
        return redirect(url_for("admin_users_page"))

    users = load_txt_users(include_inactive=True)
    target = None
    for user in users:
        if int(user.get("id", -1)) == int(user_id):
            target = user
            break
    if target is None:
        flash("対象ユーザーが見つかりません。")
        return redirect(url_for("admin_users_page"))

    duplicate = [u for u in users if int(u.get("id", -1)) != int(user_id) and normalize_username(u.get("username", "")) == username]
    if duplicate:
        flash("同じIDのユーザーが既に存在します。")
        return redirect(url_for("admin_users_page"))

    role_to_save = str(target.get("role", "staff"))
    active_to_save = int(target.get("is_active", 1) or 0)

    if int(me["id"]) == int(user_id):
        if new_role != role_to_save:
            flash("自分自身の権限は変更できません。")
            return redirect(url_for("admin_users_page"))
        if new_is_active == 0:
            flash("自分自身を退会状態にすることはできません。")
            return redirect(url_for("admin_users_page"))
        active_to_save = 1
    else:
        role_to_save = new_role
        active_to_save = new_is_active

    # 管理者が0人になる変更は禁止する。
    will_have_active_admin = False
    for user in users:
        if int(user.get("id", -1)) == int(user_id):
            candidate_role = role_to_save
            candidate_active = active_to_save
        else:
            candidate_role = str(user.get("role", "staff"))
            candidate_active = int(user.get("is_active", 0) or 0)
        if candidate_role == "admin" and candidate_active:
            will_have_active_admin = True
            break
    if not will_have_active_admin:
        flash("有効な管理者が0人になる変更はできません。")
        return redirect(url_for("admin_users_page"))

    target["username"] = username
    target["display_name"] = display_name
    target["role"] = role_to_save
    target["is_active"] = active_to_save
    target["updated_at"] = now().replace(":", "-")
    if new_password:
        target["password"] = new_password
    save_txt_users(users)

    with db() as conn:
        write_operation_log(
            conn,
            "user_update",
            "staff_users.txt",
            str(user_id),
            f"username={username}, role={role_to_save}, active={active_to_save}, password_changed={bool(new_password)}",
        )

    if int(me["id"]) == int(user_id):
        session["username"] = username
        session["role"] = role_to_save
    flash("ユーザー情報を staff_users.txt に更新しました。")
    return redirect(url_for("admin_users_page"))


@app.post("/admin/users/<int:user_id>/retire")
def admin_users_retire(user_id: int):
    guard = require_admin_or_redirect("ユーザー退会は管理者権限が必要です。")
    if guard:
        return guard

    me = current_user()
    if me is None:
        flash("ログインし直してください。")
        return redirect(url_for("staff_portal"))
    if int(me["id"]) == int(user_id):
        flash("自分自身を退会させることはできません。")
        return redirect(url_for("admin_users_page"))

    users = load_txt_users(include_inactive=True)
    target = None
    for user in users:
        if int(user.get("id", -1)) == int(user_id):
            target = user
            break
    if target is None:
        flash("対象ユーザーが見つかりません。")
        return redirect(url_for("admin_users_page"))

    active_admins = sum(1 for u in users if str(u.get("role")) == "admin" and int(u.get("is_active", 0)))
    if str(target.get("role")) == "admin" and active_admins <= 1:
        flash("最後の管理者は退会できません。")
        return redirect(url_for("admin_users_page"))

    target["is_active"] = 0
    target["updated_at"] = now().replace(":", "-")
    save_txt_users(users)
    with db() as conn:
        write_operation_log(conn, "user_retire", "staff_users.txt", str(user_id), f"username={target['username']}")
    flash("ユーザーを退会状態にしました。")
    return redirect(url_for("admin_users_page"))


@app.post("/admin/users/<int:user_id>/restore")
def admin_users_restore(user_id: int):
    guard = require_admin_or_redirect("ユーザー復帰は管理者権限が必要です。")
    if guard:
        return guard

    users = load_txt_users(include_inactive=True)
    target = None
    for user in users:
        if int(user.get("id", -1)) == int(user_id):
            target = user
            break
    if target is None:
        flash("対象ユーザーが見つかりません。")
        return redirect(url_for("admin_users_page"))
    target["is_active"] = 1
    target["updated_at"] = now().replace(":", "-")
    save_txt_users(users)
    with db() as conn:
        write_operation_log(conn, "user_restore", "staff_users.txt", str(user_id), f"username={target['username']}")
    flash("ユーザーを復帰させました。")
    return redirect(url_for("admin_users_page"))


@app.route("/staff/order-cancel")
def staff_order_cancel_page():
    """店員専用：未会計注文のキャンセル画面。"""
    guard = require_staff_or_redirect()
    if guard:
        return guard

    return render_template(
        "staff_order_cancel.html",
        store_name=STORE_NAME,
        auth_required=False,
        items=active_cancellable_order_items(),
        cancel_logs=recent_order_cancel_logs(),
    )


@app.post("/staff/order-cancel/login")
def staff_order_cancel_login():
    return redirect(url_for("staff_portal"))


@app.post("/staff/order-cancel/logout")
def staff_order_cancel_logout():
    logout_user()
    flash("ログアウトしました。")
    return redirect(url_for("staff_portal"))


@app.post("/staff/order-cancel/item/<int:item_id>")
def staff_cancel_order_item(item_id: int):
    guard = require_staff_or_redirect()
    if guard:
        return guard

    manager_no = request.form.get("manager_no", "").strip() or current_username()
    reason = request.form.get("reason", "").strip()
    try:
        cancel_quantity = int(request.form.get("cancel_quantity", "0") or 0)
    except ValueError:
        cancel_quantity = 0

    if not manager_no:
        flash("キャンセルには責任者番号が必要です。")
        return redirect(url_for("staff_order_cancel_page"))

    if cancel_quantity <= 0:
        flash("キャンセル数量は1以上で入力してください。")
        return redirect(url_for("staff_order_cancel_page"))

    with db() as conn:
        migrate_existing_db(conn)
        ensure_order_cancel_tables(conn)
        item = conn.execute(
            """
            SELECT
                oi.*,
                r.table_no,
                r.status AS receipt_status
            FROM order_items oi
            JOIN receipts r ON r.receipt_no = oi.receipt_no
            WHERE oi.id = ?
            """,
            (item_id,),
        ).fetchone()

    if item is None or item["receipt_status"] != "active":
        flash("キャンセル対象の注文が見つからないか、会計済みです。")
        return redirect(url_for("staff_order_cancel_page"))

    receipt_no = item["receipt_no"]
    lock_owner = f"cancel:{manager_no}:{time.time()}"
    ok, locked_receipt = acquire_receipt_locks([receipt_no], lock_owner)
    if not ok:
        flash(f"伝票 {locked_receipt} は別端末で処理中です。数秒後に再試行してください。")
        return redirect(url_for("staff_order_cancel_page"))

    try:
        with db() as conn:
            migrate_existing_db(conn)
            ensure_order_cancel_tables(conn)
            item = conn.execute(
                """
                SELECT
                    oi.*,
                    r.table_no,
                    r.status AS receipt_status
                FROM order_items oi
                JOIN receipts r ON r.receipt_no = oi.receipt_no
                WHERE oi.id = ?
                """,
                (item_id,),
            ).fetchone()

            if item is None or item["receipt_status"] != "active":
                flash("キャンセル対象の注文が見つからないか、会計済みです。")
                return redirect(url_for("staff_order_cancel_page"))

            quantity = int(item["quantity"] or 0)
            paid_quantity = int(item["paid_quantity"] if "paid_quantity" in item.keys() and item["paid_quantity"] is not None else 0)
            cancellable_quantity = max(0, quantity - paid_quantity)

            if cancel_quantity > cancellable_quantity:
                flash(f"キャンセル可能数量を超えています。可能数量: {cancellable_quantity}個")
                return redirect(url_for("staff_order_cancel_page"))

            unit_price = int(item["unit_price"] or 0)
            canceled_amount = unit_price * cancel_quantity
            new_quantity = quantity - cancel_quantity
            new_line_total = unit_price * new_quantity

            if new_quantity <= 0:
                conn.execute("DELETE FROM order_items WHERE id = ?", (item_id,))
            else:
                conn.execute(
                    """
                    UPDATE order_items
                    SET quantity = ?,
                        line_total = ?
                    WHERE id = ?
                    """,
                    (new_quantity, new_line_total, item_id),
                )

            conn.execute(
                """
                INSERT INTO order_cancel_logs(
                    receipt_no, table_no, order_item_id, product_name, unit_price,
                    canceled_quantity, canceled_amount, manager_no, reason, canceled_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    item["receipt_no"], int(item["table_no"]), item_id, item["product_name"],
                    unit_price, cancel_quantity, canceled_amount, manager_no, reason, now(),
                ),
            )

            recalc_unpaid_subtotal_in_conn(conn, item["receipt_no"])

        flash(f"注文をキャンセルしました: {item['product_name']} × {cancel_quantity}")
    finally:
        release_receipt_locks([receipt_no], lock_owner)

    return redirect(url_for("staff_order_cancel_page"))


@app.route("/register")
def register():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    with db() as conn:
        migrate_existing_db(conn)
        receipts = conn.execute(
            """
            SELECT r.*,
                   COALESCE((
                     SELECT SUM((oi.quantity - COALESCE(oi.paid_quantity, 0)) * oi.unit_price)
                     FROM order_items oi
                     WHERE oi.receipt_no = r.receipt_no
                       AND oi.quantity > COALESCE(oi.paid_quantity, 0)
                   ), 0) AS live_subtotal
            FROM receipts r
            WHERE r.status = 'active'
              AND r.table_no <> ?
            ORDER BY r.created_at DESC
            """,
            (TEST_TABLE_NO,),
        ).fetchall()
    return render_template("register.html", store_name=STORE_NAME, receipts=receipts)



@app.route("/reissue")
def reissue_page():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    return render_template("reissue.html", store_name=STORE_NAME)


@app.get("/api/reissue/<receipt_no>")
def api_reissue_info(receipt_no: str):
    guard = require_staff_or_json()
    if guard:
        return guard
    with db() as conn:
        migrate_existing_db(conn)
        receipt = conn.execute("SELECT * FROM receipts WHERE receipt_no = ?", (receipt_no,)).fetchone()
        if receipt is None:
            return jsonify({"ok": False, "message": "レシート番号が見つかりません。"}), 404

        items = conn.execute(
            """
            SELECT product_name, unit_price, SUM(quantity) AS quantity, SUM(line_total) AS line_total
            FROM order_items
            WHERE receipt_no = ?
            GROUP BY product_name, unit_price
            ORDER BY MIN(id)
            """,
            (receipt_no,),
        ).fetchall()

    files = list_receipt_generated_files(receipt_no)
    return jsonify({
        "ok": True,
        "receipt_no": receipt_no,
        "table_no": receipt["table_no"],
        "status": receipt["status"],
        "items": [
            {
                "product_name": row["product_name"],
                "unit_price": int(row["unit_price"] or 0),
                "quantity": int(row["quantity"] or 0),
                "line_total": int(row["line_total"] or 0),
            }
            for row in items
        ],
        "files": {key: [p.name for p in value] for key, value in files.items()},
    })


@app.post("/reissue/kitchen")
def reissue_kitchen_ticket():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    receipt_no = request.form.get("receipt_no", "").strip()
    if not receipt_no:
        flash("レシート番号を入力してください。")
        return redirect(url_for("reissue_page"))

    table_no, submitted_items = build_submitted_items_from_receipt(receipt_no)
    if table_no is None or not submitted_items:
        flash("再発行できる注文データがありません。")
        return redirect(url_for("reissue_page"))

    reissue_order_ticket_pdf_original(receipt_no)
    flash("キッチン伝票を原本と同じ表示で再発行しました。")
    return redirect(url_for("reissue_page"))


@app.post("/reissue/checkout")
def reissue_checkout_receipt():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    receipt_no = request.form.get("receipt_no", "").strip()
    if not receipt_no:
        flash("レシート番号を入力してください。")
        return redirect(url_for("reissue_page"))

    with db() as conn:
        migrate_existing_db(conn)
        receipt = conn.execute("SELECT * FROM receipts WHERE receipt_no = ?", (receipt_no,)).fetchone()
        if receipt is None:
            flash("レシート番号が見つかりません。")
            return redirect(url_for("reissue_page"))

    generate_checkout_pdf(receipt_no)
    flash("会計レシートを再発行しました。")
    return redirect(url_for("reissue_page"))


@app.route("/products/status")
def product_status_page():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    if not is_staff_admin_authenticated():
        flash("商品売切れ・販売停止は管理者権限が必要です。")
        return redirect(url_for("staff_portal"))

    try:
        products = load_product_menu(include_inactive=True)
    except Exception as e:
        products = []
        flash(str(e))
    return render_template("product_status.html", store_name=STORE_NAME, products=products)


@app.post("/products/status/toggle")
def product_status_toggle():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    if not is_staff_admin_authenticated():
        abort(403)

    line_no_raw = request.form.get("line_no", "").strip()
    action = request.form.get("action", "").strip()

    try:
        line_no = int(line_no_raw)
    except ValueError:
        flash("商品行番号が不正です。")
        return redirect(url_for("product_status_page"))

    if not MENU_TXT_PATH.exists():
        flash("product_menu.txt が見つかりません。")
        return redirect(url_for("product_status_page"))

    lines = MENU_TXT_PATH.read_text(encoding="utf-8").splitlines()
    idx = line_no - 1

    if idx < 0 or idx >= len(lines):
        flash("商品行番号が範囲外です。")
        return redirect(url_for("product_status_page"))

    original = lines[idx].rstrip()

    if not original.strip() or original.strip().startswith("#"):
        flash("この行は変更できません。")
        return redirect(url_for("product_status_page"))

    clean = re.sub(r"\|(停止|販売停止|売切れ|売り切れ)\s*$", "", original).rstrip()

    if action == "soldout":
        lines[idx] = clean + "|売切れ"
        flash("売切れにしました。")
    elif action == "stop":
        lines[idx] = clean + "|停止"
        flash("販売停止にしました。")
    elif action == "active":
        lines[idx] = clean
        flash("販売中に戻しました。")
    else:
        flash("操作が不正です。")
        return redirect(url_for("product_status_page"))

    MENU_TXT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")

    errors = validate_product_menu()
    if errors:
        lines[idx] = original
        MENU_TXT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")
        flash("商品TXTの書式エラーが出たため変更を取り消しました。")
        for err in errors[:3]:
            flash(err)

    return redirect(url_for("product_status_page"))




@app.route("/admin/sales")
def admin_sales():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    """
    管理者用：売上確認・CSV/Excel出力画面。
    レジその他画面とは分離する。
    """
    date_from, date_to = get_date_range_from_request(default_today=True)
    summary = get_sales_summary(date_from, date_to)
    return render_template(
        "admin_sales.html",
        store_name=STORE_NAME,
        date_from=date_from,
        date_to=date_to,
        summary=summary,
    )


@app.route("/register/other")
def register_other():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    date_from, date_to = get_date_range_from_request(default_today=True)
    summary = get_sales_summary(date_from, date_to)
    close_date = date_to
    opening_cash = get_opening_cash(close_date)
    daily_closing = get_daily_closing(close_date)
    return render_template(
        "register_other.html",
        store_name=STORE_NAME,
        date_from=date_from,
        date_to=date_to,
        summary=summary,
        close_date=close_date,
        opening_cash=opening_cash,
        daily_closing=daily_closing,
    )



@app.post("/register/opening-cash")
def register_opening_cash():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    date_key = date_key_from_text(request.form.get("date_key"))
    opening_cash = parse_yen_input(request.form.get("opening_cash", "0"))
    manager_no = request.form.get("manager_no", "").strip() or current_username()
    note = request.form.get("note", "").strip()
    if not manager_no:
        flash("営業開始時現金の登録には担当者番号が必要です。")
        return redirect(url_for("register_other", date_from=date_key, date_to=date_key))
    upsert_opening_cash(date_key, opening_cash, manager_no, note)
    flash(f"{date_key} の営業開始時現金を {opening_cash:,}円で登録しました。")
    return redirect(url_for("register_other", date_from=date_key, date_to=date_key))


@app.post("/register/daily-close")
def register_daily_close():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    date_key = date_key_from_text(request.form.get("date_key"))
    manager_no = request.form.get("manager_no", "").strip() or current_username()
    note = request.form.get("note", "").strip()
    if not manager_no:
        flash("日次締めには担当者番号が必要です。")
        return redirect(url_for("register_other", date_from=date_key, date_to=date_key))
    summary = get_sales_summary(date_key, date_key)
    expected_cash = int(summary.get("cash_balance", 0) or 0)
    actual_cash_raw = request.form.get("actual_cash", "")
    actual_cash = parse_yen_input(actual_cash_raw) if str(actual_cash_raw).strip() else expected_cash
    pdf_path = generate_daily_close_pdf(date_key, summary, actual_cash, manager_no, note)
    save_daily_closing(date_key, summary, actual_cash, manager_no, note, pdf_path)
    flash(f"{date_key} の日次締めPDFを作成しました。")
    return send_file(pdf_path, mimetype="application/pdf", as_attachment=False, download_name=Path(pdf_path).name)


@app.get("/daily-close/<date_key>/pdf")
def daily_close_pdf(date_key: str):
    guard = require_staff_or_redirect()
    if guard:
        return guard
    date_key = date_key_from_text(date_key)
    closing = get_daily_closing(date_key)
    pdf_path = closing.get("pdf_path", "") if closing else ""
    pdf_file = BASE_DIR / pdf_path if pdf_path else DAILY_REPORT_DIR / daily_report_filename(date_key)
    if not pdf_file.exists():
        summary = get_sales_summary(date_key, date_key)
        actual_cash = int(summary.get("cash_balance", 0) or 0)
        pdf_file = Path(generate_daily_close_pdf(date_key, summary, actual_cash, current_username() or "-", "再出力"))
        save_daily_closing(date_key, summary, actual_cash, current_username() or "-", "再出力", str(pdf_file))
    return send_file(pdf_file, mimetype="application/pdf", as_attachment=False, download_name=pdf_file.name)


@app.post("/register/refund")
def register_refund():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    receipt_no = request.form.get("receipt_no", "").strip()
    manager_no = request.form.get("manager_no", "").strip()
    reason = request.form.get("reason", "").strip()
    if not receipt_no or not manager_no:
        flash("返金には伝票番号と責任者番号が必要です。")
        return redirect(url_for("register_other"))

    lock_owner = f"refund:{manager_no}:{time.time()}"
    ok, locked = acquire_receipt_locks([receipt_no], lock_owner)
    if not ok:
        flash(f"伝票 {locked} は別端末で処理中です。")
        return redirect(url_for("register_other"))
    try:
        with db() as conn:
            migrate_existing_db(conn)
            receipt = conn.execute("SELECT * FROM receipts WHERE receipt_no = ?", (receipt_no,)).fetchone()
            if receipt is None:
                flash("伝票番号が見つかりません。")
                return redirect(url_for("register_other"))

            # 返金後は再会計できるよう active に戻す。
            # その状態で再度返金ボタンを押すと二重返金になるため、再会計前の再返金は止める。
            if receipt["status"] == "active":
                if receipt["refunded_at"]:
                    flash("この伝票は返金済みで、現在は再会計待ちです。再会計後でなければ再返金できません。")
                else:
                    flash("返金できるのは会計済み伝票のみです。")
                return redirect(url_for("register_other"))

            if receipt["status"] != "paid":
                flash("返金できるのは会計済み伝票のみです。")
                return redirect(url_for("register_other"))

            amount = int(receipt["total_amount"] or 0)
            if amount <= 0:
                # paid後にtotal_amountが0化されている場合に備え、最新会計ログから金額を取得する。
                row = conn.execute(
                    """
                    SELECT total_amount
                    FROM register_logs
                    WHERE receipt_no = ? AND checkout_type IN ('通常会計', '個別会計', '合算会計')
                    ORDER BY created_at DESC
                    LIMIT 1
                    """,
                    (receipt_no,),
                ).fetchone()
                amount = abs(int(row["total_amount"] or 0)) if row else 0

            if amount <= 0:
                flash("返金対象金額を取得できませんでした。")
                return redirect(url_for("register_other"))

            cur = conn.execute(
                """
                INSERT INTO register_logs(receipt_no, manager_no, checkout_type, payment_method, subtotal,
                    coupon_amount, discount_rate, discount_amount, total_amount, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (receipt_no, manager_no, "返金", receipt["payment_method"] or "返金", 0, 0, 0, 0, -amount, now()),
            )
            conn.execute(
                """
                INSERT INTO register_payments(register_log_id, receipt_no, payment_method, amount, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (cur.lastrowid, receipt_no, "返金", -amount, now()),
            )

            # 再会計できるよう、伝票はactiveに戻し、商品は未会計状態に戻す。
            conn.execute(
                """
                UPDATE order_items
                SET paid_quantity = 0
                WHERE receipt_no = ?
                """,
                (receipt_no,),
            )

            subtotal_row = conn.execute(
                """
                SELECT COALESCE(SUM(line_total), 0) AS subtotal
                FROM order_items
                WHERE receipt_no = ?
                """,
                (receipt_no,),
            ).fetchone()
            restored_subtotal = int(subtotal_row["subtotal"] or 0)

            conn.execute(
                """
                UPDATE receipts
                SET status = 'active',
                    subtotal = ?,
                    coupon_amount = 0,
                    discount_rate = 0,
                    discount_amount = 0,
                    total_amount = ?,
                    paid_at = NULL,
                    voided_at = NULL,
                    refunded_at = ?,
                    refund_reason = ?,
                    checkout_type = NULL,
                    payment_method = NULL,
                    manager_no = NULL
                WHERE receipt_no = ?
                """,
                (restored_subtotal, restored_subtotal, now(), reason, receipt_no),
            )

            table = conn.execute("SELECT * FROM tables WHERE table_no = ?", (receipt["table_no"],)).fetchone()
            already_gray = table is not None and table["status"] == "occupied"

            # 赤(paid)・緑(available)・その他状態でも灰色(occupied)へ戻す。
            conn.execute(
                """
                UPDATE tables
                SET status = 'occupied', current_receipt_no = ?
                WHERE table_no = ?
                """,
                (receipt_no, receipt["table_no"]),
            )

        if already_gray:
            flash("返金処理を完了しました。テーブルは既に灰色マスでした。")
        else:
            flash("返金処理を完了しました。テーブル状態を灰色マスへ戻しました。再会計できます。")
    finally:
        release_receipt_locks([receipt_no], lock_owner)
    return redirect(url_for("register_other"))



def parse_yen_input(value: str | int | float) -> int:
    """金額入力用。上限は設けず、カンマ・円記号・全角数字も受け付ける。"""
    s = str(value or "").strip()
    trans = str.maketrans("０１２３４５６７８９，．＋－￥円 ", "0123456789,.+-¥円 ")
    s = s.translate(trans)
    s = s.replace(",", "").replace("¥", "").replace("￥", "").replace("円", "").strip()
    if not s:
        return 0
    if not re.fullmatch(r"\+?\d+", s):
        return 0
    return int(s.lstrip("+") or "0")

@app.post("/register/cash-drawer")
def cash_drawer_action():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    action_type = request.form.get("action_type", "").strip()
    manager_no = request.form.get("manager_no", "").strip()
    note = request.form.get("note", "").strip()
    amount = parse_yen_input(request.form.get("amount", "0"))
    if action_type not in ("入金", "回収") or amount <= 0 or not manager_no:
        flash("レジ金処理には、区分・金額・責任者番号・1円以上の金額が必要です。金額に上限はありません。")
        return redirect(url_for("register_other"))
    with db() as conn:
        migrate_existing_db(conn)
        conn.execute(
            """
            INSERT INTO cash_drawer_logs(action_type, amount, manager_no, note, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (action_type, amount, manager_no, note, now()),
        )
    flash(f"レジ金{action_type}を記録しました。")
    return redirect(url_for("register_other"))


@app.get("/sales/export.csv")
def sales_export_csv():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    date_from, date_to = get_date_range_from_request(default_today=True)
    start = f"{date_from} 00:00:00"
    end = f"{date_to} 23:59:59"
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["created_at", "receipt_no", "manager_no", "checkout_type", "payment_method", "subtotal", "coupon", "discount_rate", "discount", "total"])
    with db() as conn:
        migrate_existing_db(conn)
        rows = conn.execute(
            """
            SELECT * FROM register_logs
            WHERE created_at BETWEEN ? AND ?
            ORDER BY created_at, id
            """,
            (start, end),
        ).fetchall()
    for r in rows:
        writer.writerow([r["created_at"], r["receipt_no"], r["manager_no"], r["checkout_type"], r["payment_method"], r["subtotal"], r["coupon_amount"], r["discount_rate"], r["discount_amount"], r["total_amount"]])
    bio = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    return send_file(bio, mimetype="text/csv", as_attachment=True, download_name=f"sales_{date_from}_{date_to}.csv")


@app.get("/sales/checkout_receipts.csv")
def checkout_receipts_export_csv():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    date_from, date_to = get_date_range_from_request(default_today=False)
    start = f"{date_from} 00:00:00"
    end = f"{date_to} 23:59:59"
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["created_at", "receipt_no", "table_no", "status", "manager_no", "checkout_type", "payment_method", "total_amount", "paid_at", "refunded_at", "refund_reason", "checkout_pdf_path"])
    with db() as conn:
        migrate_existing_db(conn)
        rows = conn.execute(
            """
            SELECT * FROM receipts
            WHERE created_at BETWEEN ? AND ?
              AND status IN ('paid', 'refunded')
            ORDER BY created_at, receipt_no
            """,
            (start, end),
        ).fetchall()
    for r in rows:
        writer.writerow([r["created_at"], r["receipt_no"], r["table_no"], r["status"], r["manager_no"], r["checkout_type"], r["payment_method"], r["total_amount"], r["paid_at"], r["refunded_at"] if "refunded_at" in r.keys() else "", r["refund_reason"] if "refund_reason" in r.keys() else "", r["checkout_pdf_path"]])
    bio = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    return send_file(bio, mimetype="text/csv", as_attachment=True, download_name=f"checkout_receipts_{date_from}_{date_to}.csv")


@app.get("/sales/hourly_genre.xlsx")
def hourly_genre_excel():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except Exception:
        flash("Excel出力には openpyxl が必要です。pip install -r requirements.txt を実行してください。")
        return redirect(url_for("register_other"))

    date_from, date_to = get_date_range_from_request(default_today=True)
    start = f"{date_from} 00:00:00"
    end = f"{date_to} 23:59:59"
    mapping = menu_category_map()

    with db() as conn:
        migrate_existing_db(conn)
        rows = conn.execute(
            """
            SELECT oi.product_name, oi.line_total, oi.quantity, oi.created_at
            FROM order_items oi
            JOIN receipts r ON r.receipt_no = oi.receipt_no
            WHERE oi.created_at BETWEEN ? AND ?
              AND r.status IN ('paid', 'refunded')
            ORDER BY oi.created_at
            """,
            (start, end),
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "RawData"
    ws.append(["日時", "時", "ジャンル", "商品名", "数量", "金額"])
    for row in rows:
        dt = row["created_at"]
        hour = int(str(dt)[11:13]) if dt else 0
        category = infer_category(row["product_name"], mapping)
        ws.append([dt, hour, category, row["product_name"], int(row["quantity"] or 0), int(row["line_total"] or 0)])

    summary = wb.create_sheet("時間帯ジャンル売上")
    categories = sorted({infer_category(r["product_name"], mapping) for r in rows}) or ["未分類"]
    summary.append(["時"] + categories + ["合計"])
    for hour in range(24):
        row_values = [hour]
        for cat in categories:
            # Excel関数でRawDataを参照する。
            row_idx = summary.max_row + 1
            col_letter = chr(65 + len(row_values))
            formula = f'=SUMIFS(RawData!$F:$F,RawData!$B:$B,$A{row_idx},RawData!$C:$C,{col_letter}$1)'
            row_values.append(formula)
        row_values.append(f"=SUM(B{summary.max_row+1}:{chr(65+len(categories))}{summary.max_row+1})")
        summary.append(row_values)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    for wsx in [ws, summary]:
        for cell in wsx[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in wsx.iter_rows():
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for col in wsx.columns:
            wsx.column_dimensions[col[0].column_letter].width = min(28, max(10, max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2))
    for row in summary.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = '#,##0'

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "時間帯別 商品ジャンル売上"
    chart.y_axis.title = "売上金額"
    chart.x_axis.title = "時"
    data = Reference(summary, min_col=2, max_col=1+len(categories), min_row=1, max_row=25)
    cats = Reference(summary, min_col=1, min_row=2, max_row=25)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 26
    summary.add_chart(chart, "A28")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"hourly_genre_sales_{date_from}_{date_to}.xlsx")

@app.post("/register/pay")
def pay():
    guard = require_staff_or_redirect()
    if guard:
        return guard
    # 複数端末・同時操作対策：会計処理中は対象伝票を短時間ロックする。
    _lock_data = request.form
    _lock_receipt_no = _lock_data.get("receipt_no", "").strip()
    _lock_checkout_type = _lock_data.get("checkout_type", "通常会計").strip() or "通常会計"
    _lock_combined = _lock_data.get("combined_receipts", "").strip()
    _lock_manager = _lock_data.get("manager_no", "").strip() or request.remote_addr or "unknown"
    _lock_receipts = receipt_numbers_for_lock(_lock_receipt_no, _lock_checkout_type, _lock_combined)
    lock_owner = f"pay:{_lock_manager}:{time.time()}"
    ok, locked_receipt = acquire_receipt_locks(_lock_receipts, lock_owner)
    if not ok:
        flash(f"伝票 {locked_receipt} は別端末で会計処理中です。数秒後に再試行してください。")
        return redirect(url_for("register"))
    try:
        data = request.form
        receipt_no = data.get("receipt_no", "").strip()
        manager_no = data.get("manager_no", "").strip()
        checkout_type = data.get("checkout_type", "通常会計").strip() or "通常会計"
        payment_method = data.get("payment_method", "").strip()
        payment_breakdown = parse_payment_breakdown(data.get("payment_breakdown", ""))
        payment_method_summary = summarize_payment_methods(payment_breakdown)
        combined_receipts_raw = data.get("combined_receipts", "").strip()

        if not receipt_no and combined_receipts_raw:
            receipt_no = combined_receipts_raw.split(",")[0].strip()

        if not receipt_no or not manager_no:
            flash("レシート番号、責任者番号は必須です。")
            return redirect(url_for("register"))

        if not payment_breakdown and not payment_method:
            flash("支払方法と使用金額を入力してください。")
            return redirect(url_for("register"))

        if not payment_method and payment_breakdown:
            payment_method = payment_method_summary

        coupon_amount = parse_yen_input(data.get("coupon_amount", "0"))
        try:
            discount_rate = int(data.get("discount_rate", "0") or 0)
        except ValueError:
            flash("割引率が不正です。")
            return redirect(url_for("register"))

        if discount_rate not in (0, 5, 10, 20):
            flash("割引率が不正です。")
            return redirect(url_for("register"))

        # 個別会計：選択商品・数量だけ会計済みにする
        if checkout_type == "個別会計":
            import json
            try:
                selected_items = json.loads(data.get("selected_items", "") or "[]")
            except json.JSONDecodeError:
                selected_items = []

            if not selected_items:
                flash("個別会計の商品が選択されていません。")
                return redirect(url_for("register"))

            subtotal = 0
            normalized = []

            with db() as conn:
                migrate_existing_db(conn)
                receipt = conn.execute(
                    "SELECT * FROM receipts WHERE receipt_no = ? AND status = 'active'",
                    (receipt_no,),
                ).fetchone()
                if receipt is None:
                    flash("有効なレシート番号が見つかりません。")
                    return redirect(url_for("register"))

                for selected in selected_items:
                    try:
                        item_id = int(selected.get("id"))
                        qty = int(selected.get("quantity"))
                    except (TypeError, ValueError):
                        continue

                    if qty <= 0:
                        continue

                    item = conn.execute(
                        """
                        SELECT id, product_name, unit_price, quantity,
                               COALESCE(paid_quantity, 0) AS paid_quantity
                        FROM order_items
                        WHERE id = ? AND receipt_no = ?
                        """,
                        (item_id, receipt_no),
                    ).fetchone()
                    if item is None:
                        continue

                    remain_qty = max(0, int(item["quantity"] or 0) - int(item["paid_quantity"] or 0))
                    pay_qty = min(qty, remain_qty)
                    if pay_qty <= 0:
                        continue

                    line_total = int(item["unit_price"] or 0) * pay_qty
                    subtotal += line_total
                    normalized.append({
                        "id": item_id,
                        "product_name": item["product_name"],
                        "unit_price": int(item["unit_price"] or 0),
                        "quantity": pay_qty,
                        "line_total": line_total,
                    })

                if not normalized:
                    flash("個別会計できる商品がありません。")
                    return redirect(url_for("register"))

                discount_amount = int(subtotal * discount_rate / 100)
                total = max(0, subtotal - coupon_amount - discount_amount)

                if payment_breakdown:
                    paid_amount = sum(int(p["amount"]) for p in payment_breakdown)
                    if paid_amount != total:
                        flash(f"支払金額の合計が会計金額と一致しません。支払合計:{paid_amount}円 / 会計金額:{total}円")
                        return redirect(url_for("register"))
                    payment_method = payment_method_summary

                for item in normalized:
                    conn.execute(
                        """
                        UPDATE order_items
                        SET paid_quantity = COALESCE(paid_quantity, 0) + ?,
                            paid_at = ?
                        WHERE id = ?
                        """,
                        (item["quantity"], now(), item["id"]),
                    )

                cur = conn.execute("""
                    INSERT INTO register_logs(
                        receipt_no, manager_no, checkout_type, payment_method, subtotal,
                        coupon_amount, discount_rate, discount_amount, total_amount, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    receipt_no, manager_no, checkout_type, payment_method, subtotal,
                    coupon_amount, discount_rate, discount_amount, total, now()
                ))
                save_register_payments(conn, cur.lastrowid, receipt_no, payment_breakdown or [{"method": payment_method, "amount": total}])

                fully_paid = mark_receipt_paid_if_complete(conn, receipt_no, manager_no, payment_method)

            # 個別会計は1回ごとに、その会計対象だけのレシートPDFを発行する。
            generate_checkout_pdf_for_items(
                receipt_no=receipt_no,
                paid_items=normalized,
                subtotal=subtotal,
                coupon_amount=coupon_amount,
                discount_rate=discount_rate,
                discount_amount=discount_amount,
                total=total,
                manager_no=manager_no,
                label="個別会計",
            )

            if not fully_paid:
                calc_unpaid_subtotal(receipt_no)

            flash("個別会計を完了しました。未会計分は伝票に残っています。")
            return redirect(url_for("register"))

        # 合算会計：複数伝票の未会計分をまとめて会計
        if checkout_type == "合算会計":
            import re
            receipt_nos = [x.strip() for x in re.split(r"[,\n\s]+", combined_receipts_raw) if x.strip()]
            if receipt_no and receipt_no not in receipt_nos:
                receipt_nos.insert(0, receipt_no)
            receipt_nos = list(dict.fromkeys(receipt_nos))

            if len(receipt_nos) < 2:
                flash("合算会計は複数のレシート番号を指定してください。")
                return redirect(url_for("register"))

            subtotal = 0
            paid_receipts = []

            with db() as conn:
                migrate_existing_db(conn)
                for rn in receipt_nos:
                    receipt = conn.execute(
                        "SELECT * FROM receipts WHERE receipt_no = ? AND status = 'active'",
                        (rn,),
                    ).fetchone()
                    if receipt is None:
                        continue

                    rows = conn.execute(
                        """
                        SELECT id, unit_price, quantity, COALESCE(paid_quantity, 0) AS paid_quantity
                        FROM order_items
                        WHERE receipt_no = ?
                          AND quantity > COALESCE(paid_quantity, 0)
                        """,
                        (rn,),
                    ).fetchall()

                    rn_subtotal = 0
                    for item in rows:
                        remain_qty = max(0, int(item["quantity"] or 0) - int(item["paid_quantity"] or 0))
                        rn_subtotal += int(item["unit_price"] or 0) * remain_qty
                        conn.execute(
                            """
                            UPDATE order_items
                            SET paid_quantity = quantity,
                                paid_at = ?
                            WHERE id = ?
                            """,
                            (now(), item["id"]),
                        )

                    if rn_subtotal > 0:
                        subtotal += rn_subtotal
                        paid_receipts.append(rn)

                if subtotal <= 0:
                    flash("合算会計できる未会計商品がありません。")
                    return redirect(url_for("register"))

                discount_amount = int(subtotal * discount_rate / 100)
                total = max(0, subtotal - coupon_amount - discount_amount)

                if payment_breakdown:
                    paid_amount = sum(int(p["amount"]) for p in payment_breakdown)
                    if paid_amount != total:
                        flash(f"支払金額の合計が会計金額と一致しません。支払合計:{paid_amount}円 / 会計金額:{total}円")
                        return redirect(url_for("register"))
                    payment_method = payment_method_summary

                cur = conn.execute("""
                    INSERT INTO register_logs(
                        receipt_no, manager_no, checkout_type, payment_method, subtotal,
                        coupon_amount, discount_rate, discount_amount, total_amount, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    ",".join(paid_receipts), manager_no, checkout_type, payment_method, subtotal,
                    coupon_amount, discount_rate, discount_amount, total, now()
                ))
                save_register_payments(conn, cur.lastrowid, ",".join(paid_receipts), payment_breakdown or [{"method": payment_method, "amount": total}])

                for rn in paid_receipts:
                    mark_receipt_paid_if_complete(conn, rn, manager_no, payment_method)

            for rn in paid_receipts:
                generate_checkout_pdf(rn)

            flash("合算会計を完了しました。")
            return redirect(url_for("register"))

        # 通常会計：伝票内の未会計分をすべて会計
        with db() as conn:
            migrate_existing_db(conn)
            receipt = conn.execute(
                "SELECT * FROM receipts WHERE receipt_no = ? AND status = 'active'",
                (receipt_no,),
            ).fetchone()
            if receipt is None:
                flash("有効なレシート番号が見つかりません。")
                return redirect(url_for("register"))

            rows = conn.execute(
                """
                SELECT id, unit_price, quantity, COALESCE(paid_quantity, 0) AS paid_quantity
                FROM order_items
                WHERE receipt_no = ?
                  AND quantity > COALESCE(paid_quantity, 0)
                """,
                (receipt_no,),
            ).fetchall()

            subtotal = 0
            paid_items_now = []
            for item in rows:
                remain_qty = max(0, int(item["quantity"] or 0) - int(item["paid_quantity"] or 0))
                line_total = int(item["unit_price"] or 0) * remain_qty
                subtotal += line_total
                if remain_qty > 0:
                    product_row = conn.execute("SELECT product_name FROM order_items WHERE id = ?", (item["id"],)).fetchone()
                    paid_items_now.append({
                        "product_name": product_row["product_name"] if product_row else "",
                        "quantity": remain_qty,
                        "line_total": line_total,
                    })
                conn.execute(
                    """
                    UPDATE order_items
                    SET paid_quantity = quantity,
                        paid_at = ?
                    WHERE id = ?
                    """,
                    (now(), item["id"]),
                )

            if subtotal <= 0:
                flash("未会計の商品がありません。")
                return redirect(url_for("register"))

            discount_amount = int(subtotal * discount_rate / 100)
            total = max(0, subtotal - coupon_amount - discount_amount)

            if payment_breakdown:
                paid_amount = sum(int(p["amount"]) for p in payment_breakdown)
                if paid_amount != total:
                    flash(f"支払金額の合計が会計金額と一致しません。支払合計:{paid_amount}円 / 会計金額:{total}円")
                    return redirect(url_for("register"))
                payment_method = payment_method_summary

            conn.execute("""
                UPDATE receipts
                SET status = 'paid', subtotal = ?, coupon_amount = ?, discount_rate = ?,
                    discount_amount = ?, total_amount = ?, checkout_type = ?,
                    payment_method = ?, manager_no = ?, paid_at = ?, voided_at = ?
                WHERE receipt_no = ?
            """, (
                subtotal, coupon_amount, discount_rate, discount_amount, total,
                checkout_type, payment_method, manager_no, now(), now(), receipt_no
            ))
            conn.execute("UPDATE tables SET status = 'paid' WHERE table_no = ?", (receipt["table_no"],))
            cur = conn.execute("""
                INSERT INTO register_logs(
                    receipt_no, manager_no, checkout_type, payment_method, subtotal,
                    coupon_amount, discount_rate, discount_amount, total_amount, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                receipt_no, manager_no, checkout_type, payment_method, subtotal,
                coupon_amount, discount_rate, discount_amount, total, now()
            ))
            save_register_payments(conn, cur.lastrowid, receipt_no, payment_breakdown or [{"method": payment_method, "amount": total}])

        generate_checkout_pdf_for_items(
            receipt_no=receipt_no,
            paid_items=paid_items_now,
            subtotal=subtotal,
            coupon_amount=coupon_amount,
            discount_rate=discount_rate,
            discount_amount=discount_amount,
            total=total,
            manager_no=manager_no,
            label="会計",
        )
        flash("会計を完了しました。")
        return redirect(url_for("register"))




    finally:
        release_receipt_locks(_lock_receipts, lock_owner)


@app.route("/receipt/<receipt_no>/checkout-pdf")
def checkout_receipt_pdf(receipt_no: str):
    guard = require_staff_or_redirect()
    if guard:
        return guard
    """手動確認用。通常運用では会計確定時にバックエンド保存のみ行う。"""
    with db() as conn:
        migrate_existing_db(conn)
        receipt = conn.execute("SELECT * FROM receipts WHERE receipt_no = ?", (receipt_no,)).fetchone()
    if receipt is None:
        abort(404)

    pdf_path = receipt["checkout_pdf_path"] if "checkout_pdf_path" in receipt.keys() else None
    if pdf_path:
        pdf_file = BASE_DIR / pdf_path
    else:
        pdf_file = Path(generate_checkout_pdf(receipt_no))

    if not pdf_file.exists():
        pdf_file = Path(generate_checkout_pdf(receipt_no))

    return send_file(pdf_file, mimetype="application/pdf", as_attachment=False, download_name=pdf_file.name)



@app.post("/mobile/<token>/order")
def mobile_order_submit(token: str):
    """互換用。旧URLからもcart_json形式で注文できるようにする。"""
    with db() as conn:
        migrate_existing_db(conn)
        receipt = conn.execute("SELECT * FROM receipts WHERE token = ?", (token,)).fetchone()

    if receipt is None or receipt["status"] != "active":
        return render_template("mobile_disabled.html", store_name=STORE_NAME), 403

    cart_raw = request.form.get("cart_json", "[]")
    try:
        cart = json.loads(cart_raw)
    except json.JSONDecodeError:
        cart = []

    if not isinstance(cart, list):
        cart = []

    submitted_items = []
    cart_errors = []
    lock_owner = f"order:{receipt['receipt_no']}:{request.remote_addr or 'unknown'}:{time.time()}"
    ok, locked_receipt = acquire_receipt_locks([receipt["receipt_no"]], lock_owner, ttl_seconds=45)
    if not ok:
        flash("ただいま店員が会計・取消処理中です。追加注文は店員にお声がけください。")
        return redirect(url_for("mobile_order", token=token))

    try:
        with db() as conn:
            migrate_existing_db(conn)
            fresh_receipt = conn.execute(
                "SELECT * FROM receipts WHERE receipt_no = ?",
                (receipt["receipt_no"],),
            ).fetchone()
            if fresh_receipt is None or fresh_receipt["status"] != "active":
                return render_template("mobile_disabled.html", store_name=STORE_NAME), 403

            built_items, cart_errors = build_order_items_from_cart_with_errors(cart)
            submitted_items = insert_built_order_items(conn, receipt, built_items)

            if submitted_items:
                recalc_receipt_total(conn, receipt["receipt_no"])

        if submitted_items:
            generate_order_ticket_pdf(
                receipt_no=receipt["receipt_no"],
                table_no=receipt["table_no"],
                submitted_items=submitted_items,
            )
            flash("注文を送信しました。")
        else:
            if cart_errors:
                for error in cart_errors:
                    flash(error)
            else:
                flash("商品を1つ以上選択してください。")
    finally:
        release_receipt_locks([receipt["receipt_no"]], lock_owner)

    return redirect(url_for("mobile_order", token=token))


@app.route("/mobile/<token>/complete/<int:order_id>")
def mobile_order_complete(token: str, order_id: int):
    with db() as conn:
        receipt = conn.execute("SELECT * FROM receipts WHERE token = ?", (token,)).fetchone()

        if receipt is None or receipt["status"] != "active":
            return render_template("mobile_disabled.html", store_name=STORE_NAME), 403

        order = conn.execute(
            "SELECT * FROM orders WHERE id = ? AND receipt_no = ?",
            (order_id, receipt["receipt_no"]),
        ).fetchone()

        if order is None:
            abort(404)

        items = conn.execute(
            "SELECT * FROM order_items WHERE order_id = ? ORDER BY id",
            (order_id,),
        ).fetchall()

        subtotal = recalc_receipt_total(conn, receipt["receipt_no"])

    return render_template(
        "mobile_complete.html",
        store_name=STORE_NAME,
        receipt=receipt,
        order=order,
        items=items,
        subtotal=subtotal,
    )



if __name__ == "__main__":
    init_db()

    lan_base = get_base_url()

    print("=" * 60)
    print("スタッフメニューから各機能へ移動してください。")
    print(f"PC            : http://127.0.0.1:{APP_PORT}/staff")
    print(f"同じWi-Fi端末 : {lan_base}/staff")
    print("-" * 60)
    print("注意: スマホはPCと同じWi-Fiに接続してください。")
    print("注意: Windows Defender ファイアウォールで Python の通信を許可してください。")
    print("運用時推奨: python run_waitress.py で起動してください。")
    print("=" * 60)

    app.run(debug=False, threaded=True, host="0.0.0.0", port=APP_PORT)
